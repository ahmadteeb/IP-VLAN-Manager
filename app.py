from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, send_file
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.exceptions import abort
from waitress import serve
from models.models import db, User, IP, VLAN, ActivityLog, Router, Interface, Site, StatusType, UserRole, Vendor, PasswordState, Technology
from config import Config
from datetime import datetime
import csv
import math
from functools import wraps
import ipaddress
import io
from sqlalchemy.exc import IntegrityError

app = Flask(__name__)
app.config.from_object(Config)

# Initialize extensions
db.init_app(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Please log in to access this page.'

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

def admin_required(f):
    """Decorator to require admin role"""
    @wraps(f)
    @login_required
    def decorated_function(*args, **kwargs):
        if not current_user.is_admin():
            abort(403)
        return f(*args, **kwargs)
    return decorated_function

def write_access_required(f):
    """Decorator to disallow actions for read-only users"""
    @wraps(f)
    @login_required
    def decorated_function(*args, **kwargs):
        try:
            if hasattr(current_user, 'is_read_only') and current_user.is_read_only():
                abort(403)
        except Exception as e:
            app.logger.error(f'Error in write_access_required: {str(e)}', exc_info=True)
            abort(403)
        return f(*args, **kwargs)
    return decorated_function
def log_activity(action, resource_type, resource_id, resource_value, site_name=None, router=None, interface=None):
    """Helper function to log activities"""
    log = ActivityLog(
        user_id=current_user.id,
        user_username=current_user.username,
        action=action,
        resource_type=resource_type,
        resource_id=resource_id,
        resource_value=resource_value,
        site_name=site_name,
        router=router,
        interface=interface
    )
    db.session.add(log)
    db.session.commit()

def parse_technology(tech_str):
    """Validate technology by name against the Technologies table.
    Returns canonical name if exists, otherwise None."""
    if not tech_str:
        return None
    try:
        # Case-insensitive lookup by name
        tech = Technology.query.filter(db.func.lower(Technology.name) == tech_str.strip().lower()).first()
        return tech.name if tech else None
    except Exception:
        return None

# Routes
@app.route('/')
@login_required
def index():
    return redirect(url_for('dashboard'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        # If user still must change password, redirect to change password page
        try:
            state = PasswordState.query.filter_by(user_id=current_user.id).first()
            if state and state.must_change:
                return redirect(url_for('change_password'))
        except Exception as e:
            app.logger.error(f'Error checking password state on existing session: {str(e)}', exc_info=True)
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        remember = bool(request.form.get('remember'))
        
        user = User.query.filter_by(username=username).first()
        
        if user and user.check_password(password):
            login_user(user, remember=remember)
            # Ensure a PasswordState exists
            try:
                state = PasswordState.query.filter_by(user_id=user.id).first()
                if not state:
                    state = PasswordState(user_id=user.id, must_change=True)
                    db.session.add(state)
                    db.session.commit()
            except Exception as e:
                app.logger.error(f'Error ensuring password state for user {user.username}: {str(e)}', exc_info=True)
            # If required, force password change
            try:
                state = PasswordState.query.filter_by(user_id=user.id).first()
                if state and state.must_change:
                    return redirect(url_for('change_password'))
            except Exception as e:
                app.logger.error(f'Error checking password state post login: {str(e)}', exc_info=True)
            next_page = request.args.get('next')
            return redirect(next_page) if next_page else redirect(url_for('dashboard'))
        else:
            flash('Invalid username or password', 'error')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    # Get statistics
    total_routers = Router.query.count()
    total_interfaces = Interface.query.count()
    total_sites = Site.query.count()
    total_ips = IP.query.count()
    assigned_ips = IP.query.filter_by(status=StatusType.ASSIGNED).count()
    free_ips = total_ips - assigned_ips
    total_vlans = VLAN.query.count()
    assigned_vlans = VLAN.query.filter_by(status=StatusType.ASSIGNED).count()
    free_vlans = total_vlans - assigned_vlans
    
    # Stats by technology
    site_stats_by_tech = {}
    ip_stats_by_tech = {}
    vlan_stats_by_tech = {}
    for t in Technology.query.all():
        site_count = Site.query.filter_by(technology_type=t.name).count()
        ip_count = IP.query.filter_by(type=t.name).count()
        ip_assigned = IP.query.filter_by(type=t.name, status=StatusType.ASSIGNED).count()
        vlan_count = VLAN.query.filter_by(type=t.name).count()
        vlan_assigned = VLAN.query.filter_by(type=t.name, status=StatusType.ASSIGNED).count()
        
        site_stats_by_tech[t.name] = site_count
        ip_stats_by_tech[t.name] = {
            'total': ip_count,
            'assigned': ip_assigned,
            'free': ip_count - ip_assigned
        }
        vlan_stats_by_tech[t.name] = {
            'total': vlan_count,
            'assigned': vlan_assigned,
            'free': vlan_count - vlan_assigned
        }
    
    # Recent activities
    recent_activities = ActivityLog.query.order_by(ActivityLog.timestamp.desc()).limit(10).all()
    
    return render_template('dashboard.html',
                         total_routers=total_routers,
                         total_interfaces=total_interfaces,
                         total_sites=total_sites,
                         total_ips=total_ips,
                         assigned_ips=assigned_ips,
                         free_ips=free_ips,
                         total_vlans=total_vlans,
                         assigned_vlans=assigned_vlans,
                         free_vlans=free_vlans,
                         site_stats_by_tech=site_stats_by_tech,
                         ip_stats_by_tech=ip_stats_by_tech,
                         vlan_stats_by_tech=vlan_stats_by_tech,
                         recent_activities=recent_activities)

# Technologies API
@app.route('/api/technologies', methods=['GET'])
@login_required
def api_get_technologies():
    """Return list of technologies from DB."""
    try:
        search = request.args.get('search', "", type=str)
        query = Technology.query.order_by(Technology.id.asc())
        
        if search:
            query = query.filter(Technology.name.like(f'%{search}%'))
        
        technologies = query.all()
        return jsonify({
            'technologies': [t.to_dict() for t in technologies],
            'search': search
        })
    except Exception as e:
        app.logger.error(f'Error fetching technologies: {str(e)}', exc_info=True)
        return jsonify({'error': 'Failed to load technologies'}), 500

@app.route('/api/technologies', methods=['POST'])
@login_required
@admin_required
def api_create_technology():
    """Create a new technology entry"""
    try:
        data = request.json or {}
        name = (data.get('name') or '').strip()
        if not name:
            return jsonify({'error': 'Name is required'}), 400
        if Technology.query.filter_by(name=name).first():
            return jsonify({'error': 'Technology name already exists'}), 400
        tech = Technology(name=name)
        db.session.add(tech)
        db.session.commit()
        return jsonify({'message': 'Technology created', 'technology': tech.to_dict()}), 201
    except Exception as e:
        app.logger.error(f'Error creating technology: {str(e)}', exc_info=True)
        return jsonify({'error': 'Failed to create technology'}), 500

@app.route('/api/technologies/<int:tech_id>', methods=['PUT'])
@login_required
@admin_required
def api_update_technology(tech_id):
    """Update a technology entry"""
    try:
        tech = Technology.query.get_or_404(tech_id)
        data = request.json or {}
        if 'name' in data and data['name']:
            new_name = data['name'].strip()
            if Technology.query.filter(Technology.id != tech_id, Technology.name == new_name).first():
                return jsonify({'error': 'Technology name already exists'}), 400
            tech.name = new_name
        db.session.commit()
        return jsonify({'message': 'Technology updated', 'technology': tech.to_dict()})
    except Exception as e:
        app.logger.error(f'Error updating technology {tech_id}: {str(e)}', exc_info=True)
        return jsonify({'error': 'Failed to update technology'}), 500

@app.route('/api/technologies/<int:tech_id>', methods=['DELETE'])
@login_required
@admin_required
def api_delete_technology(tech_id):
    """Delete a technology entry"""
    try:
        tech = Technology.query.get_or_404(tech_id)
        db.session.delete(tech)
        db.session.commit()
        return jsonify({'message': 'Technology deleted'})
    except Exception as e:
        app.logger.error(f'Error deleting technology {tech_id}: {str(e)}', exc_info=True)
        return jsonify({'error': 'Failed to delete technology'}), 500

# Router Management Routes
@app.route('/routers')
@login_required
def routers():
    return render_template('routers.html')

@app.route('/technologies')
@login_required
def technologies():
    return render_template('technologies.html')

@app.route('/api/routers', methods=['GET'])
@login_required
def api_get_routers():
    """Get routers with pagination"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    search = request.args.get('search', "", type=str)
    
    query = Router.query.order_by(Router.name)
    if search:
        query = query.filter(Router.name.like(f'%{search}%'))

    pagination = query.paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    return jsonify({
        'routers': [r.to_dict() for r in pagination.items],
        'total': pagination.total,
        'pages': pagination.pages,
        'current_page': page,
        'search': search
    })

@app.route('/api/routers', methods=['POST'])
@login_required
@admin_required
def api_create_router():
    """Create a new router"""
    data = request.json
    name = data.get('name')
    router_ip = data.get('router_ip')
    router_type = data.get('router_type')
    description = data.get('description', '')
    
    if not all([name, router_ip, router_type]):
        return jsonify({'error': 'Router name, IP, and type are required'}), 400
    
    if Router.query.filter_by(name=name).first():
        return jsonify({'error': 'Router with this name already exists'}), 400
    
    if Router.query.filter_by(router_ip=router_ip).first():
        return jsonify({'error': 'Router with this IP already exists'}), 400
    
    router = Router(name=name, router_ip=router_ip, router_type=router_type, description=description)
    db.session.add(router)
    db.session.commit()
    
    log_activity('create_router', 'router', router.id, router.name)
    
    return jsonify({'message': 'Router created successfully', 'router': router.to_dict()}), 201

@app.route('/api/routers/<int:router_id>', methods=['PUT'])
@login_required
@admin_required
def api_update_router(router_id):
    """Update a router"""
    router = Router.query.get_or_404(router_id)
    data = request.json
    
    if 'name' in data:
        if Router.query.filter(Router.id != router_id, Router.name == data['name']).first():
            return jsonify({'error': 'Router with this name already exists'}), 400
        router.name = data['name']
    
    if 'router_ip' in data:
        if Router.query.filter(Router.id != router_id, Router.router_ip == data['router_ip']).first():
            return jsonify({'error': 'Router with this IP already exists'}), 400
        router.router_ip = data['router_ip']
    
    if 'router_type' in data:
        router.router_type = data['router_type']
    
    if 'description' in data:
        router.description = data['description']
    
    db.session.commit()
    
    return jsonify({'message': 'Router updated successfully', 'router': router.to_dict()}), 200

@app.route('/api/routers/<int:router_id>', methods=['DELETE'])
@login_required
@admin_required
def api_delete_router(router_id):
    """Delete a router"""
    router = Router.query.get_or_404(router_id)
    
    # Check if router has interfaces with sites
    for interface in router.interfaces:
        if interface.sites:
            return jsonify({'error': 'Cannot delete router with assigned sites'}), 400
    
    db.session.delete(router)
    db.session.commit()
    
    return jsonify({'message': 'Router deleted successfully'}), 200

# Interface Management Routes
@app.route('/interfaces')
@login_required
def interfaces():
    return render_template('interfaces.html')

@app.route('/api/interfaces', methods=['GET'])
@login_required
def api_get_interfaces():
    """Get interfaces with optional filtering and pagination"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    router_id = request.args.get('router_id', type=int)
    search = request.args.get('search', "", type=str)
    
    query = Interface.query
    
    if router_id:
        query = query.filter_by(router_id=router_id)
    
    if search:
        query = query.filter(Interface.name.like(f'%{search}%'))
    
    query = query.order_by(Interface.name)
    
    pagination = query.paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    return jsonify({
        'interfaces': [i.to_dict() for i in pagination.items],
        'total': pagination.total,
        'pages': pagination.pages,
        'current_page': page,
        'search': search
    })

@app.route('/api/interfaces', methods=['POST'])
@login_required
@admin_required
def api_create_interface():
    """Create a new interface"""
    data = request.json
    router_id = data.get('router_id')
    name = data.get('name')
    
    if not all([router_id, name]):
        return jsonify({'error': 'Router and Interface Name are required'}), 400
    
    router = Router.query.get_or_404(router_id)
    
    # Check if interface already exists on this router
    if Interface.query.filter_by(router_id=router_id, name=name).first():
        return jsonify({'error': 'Interface with this name already exists on this router'}), 400
    
    interface = Interface(
        router_id=router_id,
        name=name
    )
    db.session.add(interface)
    db.session.commit()
    
    log_activity('create_interface', 'interface', interface.id, interface.name,
                router=router.name, interface=interface.name)
    
    return jsonify({'message': 'Interface created successfully', 'interface': interface.to_dict()}), 201

@app.route('/api/interfaces/<int:interface_id>', methods=['DELETE'])
@login_required
@admin_required
def api_delete_interface(interface_id):
    """Delete an interface"""
    interface = Interface.query.get_or_404(interface_id)
    
    # Check if interface has assigned sites
    if interface.sites:
        return jsonify({'error': 'Cannot delete interface with assigned sites'}), 400
    
    router_name = interface.router.name
    interface_name = interface.name
    db.session.delete(interface)
    db.session.commit()
    
    return jsonify({'message': 'Interface deleted successfully'}), 200

# VLAN Management Routes (Predefined VLANs)
@app.route('/vlans')
@login_required
def vlans():
    return render_template('vlans.html')

@app.route('/api/vlans', methods=['GET'])
@login_required
def api_get_vlans():
    """Get VLANs with filtering and pagination"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    technology = request.args.get('technology')
    vendor = request.args.get('vendor')
    status = request.args.get('status')
    search = request.args.get('search', '')
    
    query = VLAN.query
    
    if technology:
        tech_name = parse_technology(technology)
        if tech_name:
            # Include both service and OM VLANs for the technology
            query = query.filter(db.or_(
                VLAN.type == tech_name,
                VLAN.type == f"{tech_name}_OM"
            ))
    if vendor:
        # Support both vendor_id and vendor string for backward compatibility
        try:
            vendor_id = int(vendor)
            query = query.filter_by(vendor_id=vendor_id)
        except (ValueError, TypeError):
            query = query.filter_by(vendor=vendor)
    if status:
        try:
            status_enum = StatusType[status]
            query = query.filter_by(status=status_enum)
        except KeyError:
            pass
    if search:
        try:
            vlan_id = int(search)
            query = query.filter(VLAN.vlan_id == vlan_id)
        except ValueError:
            # Search by site name or vendor through relationship (left join to include VLANs without sites)
            query = query.outerjoin(Site, db.or_(
                VLAN.id == Site.service_vlan_id,
                VLAN.id == Site.om_vlan_id
            )).filter(
                db.or_(
                    Site.site_name.like(f'%{search}%'),
                    Site.site_id.like(f'%{search}%'),
                    VLAN.vendor.like(f'%{search}%')
                )
            ).distinct()
    
    # For pagination, we only count service VLANs (or unpaired VLANs) since that's what we display
    # This ensures pagination matches what's actually shown in the table
    service_query = query.filter(db.or_(
        VLAN.pair_type.is_(None),
        VLAN.pair_type == 'service'
    ))
    
    service_query = service_query.order_by(VLAN.vlan_id)
    
    pagination = service_query.paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    # Return the service VLANs - their to_dict() method will include pair_vlan_number for OM VLANs
    return jsonify({
        'vlans': [v.to_dict() for v in pagination.items],
        'total': pagination.total,
        'pages': pagination.pages,
        'current_page': page
    })

@app.route('/api/vlans', methods=['POST'])
@login_required
@admin_required
def api_create_vlan():
    """Create a VLAN (pool-based, not tied to interface, can be reused)"""
    data = request.json
    vlan_id = data.get('vlan_id')
    technology = data.get('technology')
    vendor_id = data.get('vendor_id')
    vendor = data.get('vendor')  # Keep for backward compatibility
    
    if not all([vlan_id, technology]):
        return jsonify({'error': 'VLAN ID and Technology are required'}), 400
    
    if not vendor_id and not vendor:
        return jsonify({'error': 'Vendor is required'}), 400
    
    tech_name = parse_technology(technology)
    if not tech_name:
        return jsonify({'error': 'Invalid technology'}), 400
    
    # Normalize vendor
    # Get vendor object if vendor_id is provided
    vendor_obj = None
    vendor_name = None
    if vendor_id:
        vendor_obj = Vendor.query.get(vendor_id)
        if not vendor_obj:
            return jsonify({'error': 'Invalid vendor ID'}), 400
        vendor_name = vendor_obj.name
    else:
        vendor_name = vendor

    # Accept either a single VLAN ID or a range string "start-end"
    vlan_ids_to_create = []
    if isinstance(vlan_id, int):
        vlan_ids_to_create = [vlan_id]
    elif isinstance(vlan_id, str):
        vlan_id = vlan_id.strip()
        if '-' in vlan_id:
            parts = vlan_id.split('-', 1)
            try:
                start = int(parts[0].strip())
                end = int(parts[1].strip())
            except (ValueError, TypeError):
                return jsonify({'error': 'Invalid VLAN range format. Use e.g., "20-25".'}), 400
            if start > end:
                return jsonify({'error': 'Invalid VLAN range: start cannot be greater than end'}), 400
            vlan_ids_to_create = list(range(start, end + 1))
        else:
            try:
                single_id = int(vlan_id)
                vlan_ids_to_create = [single_id]
            except (ValueError, TypeError):
                return jsonify({'error': 'VLAN ID must be a number or a range like "20-25".'}), 400
    else:
        return jsonify({'error': 'VLAN ID must be a number or a range like "20-25".'}), 400

    # Validate bounds
    for vid in vlan_ids_to_create:
        if vid < 1 or vid > 4095:
            return jsonify({'error': f'VLAN ID {vid} is out of range. Must be between 1 and 4095'}), 400
    
    # Check if creating pairs
    create_pair = data.get('create_pair', False)
    om_vlan_id = data.get('om_vlan_id')  # Optional OM VLAN ID or range for pair creation
    
    # Parse OM VLAN IDs if creating pair
    om_vlan_ids_to_create = []
    if create_pair:
        if not om_vlan_id:
            return jsonify({'error': 'OM VLAN ID is required when creating a pair'}), 400
        
        # Parse OM VLAN ID (can be single or range)
        if isinstance(om_vlan_id, int):
            om_vlan_ids_to_create = [om_vlan_id]
        elif isinstance(om_vlan_id, str):
            om_vlan_id = om_vlan_id.strip()
            if '-' in om_vlan_id:
                parts = om_vlan_id.split('-', 1)
                try:
                    start = int(parts[0].strip())
                    end = int(parts[1].strip())
                except (ValueError, TypeError):
                    return jsonify({'error': 'Invalid OM VLAN range format. Use e.g., "20-25".'}), 400
                if start > end:
                    return jsonify({'error': 'Invalid OM VLAN range: start cannot be greater than end'}), 400
                om_vlan_ids_to_create = list(range(start, end + 1))
            else:
                try:
                    single_id = int(om_vlan_id)
                    om_vlan_ids_to_create = [single_id]
                except (ValueError, TypeError):
                    return jsonify({'error': 'OM VLAN ID must be a number or a range like "20-25".'}), 400
        else:
            return jsonify({'error': 'OM VLAN ID must be a number or a range like "20-25".'}), 400
        
        # Validate OM VLAN bounds
        for om_vid in om_vlan_ids_to_create:
            if om_vid < 1 or om_vid > 4095:
                return jsonify({'error': f'OM VLAN ID {om_vid} is out of range. Must be between 1 and 4095'}), 400
        
        # Check that service and OM ranges have the same count
        if len(vlan_ids_to_create) != len(om_vlan_ids_to_create):
            return jsonify({
                'error': f'Service VLAN range ({len(vlan_ids_to_create)} VLANs) and OM VLAN range ({len(om_vlan_ids_to_create)} VLANs) must have the same number of VLANs'
            }), 400
    
    # Ensure all requested VLAN IDs are unique (across service and OM)
    requested_vlan_ids = vlan_ids_to_create + om_vlan_ids_to_create
    seen_vlan_ids = set()
    duplicates_in_request = set()
    for vid in requested_vlan_ids:
        if vid in seen_vlan_ids:
            duplicates_in_request.add(vid)
        else:
            seen_vlan_ids.add(vid)
    if duplicates_in_request:
        duplicates_list = ', '.join(str(v) for v in sorted(duplicates_in_request))
        return jsonify({'error': f'Duplicate VLAN IDs in request: {duplicates_list}'}), 400

    # Check existing VLANs in database
    if seen_vlan_ids:
        existing_vlans = VLAN.query.filter(VLAN.vlan_id.in_(seen_vlan_ids)).with_entities(VLAN.vlan_id).all()
        if existing_vlans:
            existing_values = ', '.join(str(v[0]) for v in existing_vlans)
            return jsonify({'error': f'VLAN IDs already exist: {existing_values}'}), 400
    
    # Create VLANs (pool-based, reuse allowed, no uniqueness checks)
    created = []
    created_om = []
    try:
        import uuid
        for idx, vid in enumerate(vlan_ids_to_create):
            # Generate unique pair_id for each pair
            pair_id = None
            if create_pair:
                pair_id = str(uuid.uuid4())
            
            # Create service VLAN
            vlan = VLAN(
                vlan_id=vid,
                type=tech_name,
                vendor_id=vendor_obj.id if vendor_obj else None,
                vendor=vendor_name,
                status=StatusType.FREE,
                pair_id=pair_id if create_pair else None,
                pair_type='service' if create_pair else None
            )
            db.session.add(vlan)
            created.append(vlan)
            
            # Create OM VLAN if creating pair
            if create_pair:
                if idx >= len(om_vlan_ids_to_create):
                    db.session.rollback()
                    return jsonify({'error': f'Index out of range: service VLAN {vid} has no matching OM VLAN'}), 400
                om_vid = om_vlan_ids_to_create[idx]
                om_vlan = VLAN(
                    vlan_id=om_vid,
                    type=f"{tech_name}_OM",
                    vendor_id=vendor_obj.id if vendor_obj else None,
                    vendor=vendor_name,
                    status=StatusType.FREE,
                    pair_id=pair_id,
                    pair_type='om'
                )
                db.session.add(om_vlan)
                created_om.append(om_vlan)
        
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return jsonify({'error': 'One or more VLAN IDs already exist'}), 400
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Error creating VLANs: {str(e)}', exc_info=True)
        return jsonify({'error': f'Error creating VLANs: {str(e)}'}), 500

    # Log single activity summarizing the bulk creation
    if created:
        log_activity('create_vlan', 'vlan', created[0].id, 
                    f'Created {len(created)} VLAN(s)' + (f' with {len(created_om)} OM pair(s)' if created_om else ''))

    result = {
        'message': f'Created {len(created)} VLAN(s) successfully' + (' (with OM pairs)' if created_om else ''),
        'count': len(created),
        'vlans': [v.to_dict() for v in created[:10]]
    }
    if created_om:
        result['om_vlans'] = [v.to_dict() for v in created_om[:10]]
        result['count'] = len(created) + len(created_om)
    
    if len(created) == 1 and not created_om:
        result['vlan'] = created[0].to_dict()
    
    return jsonify(result), 201

@app.route('/api/vlans/<int:vlan_id>', methods=['DELETE'])
@login_required
@admin_required
def api_delete_vlan(vlan_id):
    """Delete a VLAN (cannot delete if assigned to any site)"""
    vlan = VLAN.query.get_or_404(vlan_id)
    
    # Check if VLAN is assigned to any site
    if vlan.sites:
        return jsonify({'error': 'Cannot delete VLAN that is assigned to a site'}), 400
    
    db.session.delete(vlan)
    db.session.commit()
    
    return jsonify({'message': 'VLAN deleted successfully'}), 200

# IP Management Routes
@app.route('/ips')
@login_required
def ips():
    return render_template('ips.html')

@app.route('/api/ips', methods=['GET'])
@login_required
def api_get_ips():
    """Get IPs with filtering and pagination"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        technology = request.args.get('technology')
        vendor = request.args.get('vendor')
        status = request.args.get('status')
        search = request.args.get('search', '')
        
        query = IP.query
        
        if technology:
            tech_enum = parse_technology(technology)
            if tech_enum:
                # Include both service and OM IPs for the technology
                query = query.filter(db.or_(
                    IP.type == tech_enum,
                    IP.type == f"{tech_enum}_OM"
                ))
        if vendor:
            # Support both vendor_id and vendor string for backward compatibility
            try:
                vendor_id = int(vendor)
                query = query.filter_by(vendor_id=vendor_id)
            except (ValueError, TypeError):
                query = query.filter_by(vendor=vendor)
        if status:
            try:
                status_enum = StatusType[status]
                query = query.filter_by(status=status_enum)
            except KeyError:
                pass
        if search:
            query = query.filter(IP.gateway.like(f'%{search}%'))
        
        # For pagination, we only count service IPs (or unpaired IPs) since that's what we display
        # This ensures pagination matches what's actually shown in the table
        service_query = query.filter(db.or_(
            IP.pair_type.is_(None),
            IP.pair_type == 'service'
        ))
        
        # Get all service IPs for sorting (we need to sort before pagination)
        all_service_ips = service_query.all()
        # Sort IPs numerically by IP address
        all_service_ips.sort(key=lambda ip: ipaddress.IPv4Address(ip.gateway))
        
        # Apply pagination manually after sorting
        total_count = len(all_service_ips)
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page
        paginated_service_ips = all_service_ips[start_idx:end_idx]
        
        # Return only service IPs - their to_dict() method will include pair_gateway for OM IPs
        total_pages = math.ceil(total_count / per_page) if total_count > 0 else 1
        
        app.logger.info(f'IP pagination: page={page}, per_page={per_page}, total={total_count}, pages={total_pages}')
        
        return jsonify({
            'ips': [ip.to_dict() for ip in paginated_service_ips],
            'total': total_count,
            'pages': total_pages,
            'current_page': page
        })
    except Exception as e:
        app.logger.error(f'Error in api_get_ips: {str(e)}', exc_info=True)
        return jsonify({'error': 'An error occurred while fetching IPs'}), 500

@app.route('/api/ips', methods=['POST'])
@login_required
@admin_required
def api_create_ip():
    """Create a new IP address or generate IPs from subnet"""
    data = request.json
    add_method = data.get('add_method', 'single')
    technology = data.get('technology')
    vendor_id = data.get('vendor_id')
    vendor = data.get('vendor')  # Keep for backward compatibility
    
    if not technology:
        return jsonify({'error': 'Technology is required'}), 400
    
    if not vendor_id and not vendor:
        return jsonify({'error': 'Vendor is required'}), 400
    
    tech_name = parse_technology(technology)
    if not tech_name:
        return jsonify({'error': 'Invalid technology'}), 400
    
    # Get vendor object if vendor_id is provided
    vendor_obj = None
    vendor_name = None
    if vendor_id:
        vendor_obj = Vendor.query.get(vendor_id)
        if not vendor_obj:
            return jsonify({'error': 'Invalid vendor ID'}), 400
        vendor_name = vendor_obj.name
    else:
        vendor_name = vendor
    
    if add_method == 'subnetting':
        # Handle subnetting calculator - divide base subnet and generate IPs from all subnets (with optional pair creation)
        base_subnet = data.get('base_subnet')
        num_subnets = data.get('num_subnets')
        create_pair = data.get('create_pair', False)
        om_base_subnet = data.get('om_base_subnet')  # Optional OM base subnet for pair creation
        
        if not base_subnet or not num_subnets:
            return jsonify({'error': 'Base subnet and number of subnets are required'}), 400
        
        if create_pair and not om_base_subnet:
            return jsonify({'error': 'OM base subnet is required when creating a pair'}), 400
        
        try:
            num_subnets = int(num_subnets)
            if num_subnets < 2:
                return jsonify({'error': 'Number of subnets must be at least 2'}), 400
        except (ValueError, TypeError):
            return jsonify({'error': 'Number of subnets must be a valid integer'}), 400
        
        try:
            import uuid
            
            # Helper function to divide a base network into subnets
            def divide_network(base_subnet_str, num_subnets_int):
                base_network = ipaddress.ip_network(base_subnet_str, strict=False)
                base_prefix = base_network.prefixlen
                
                # Calculate required prefix length for the number of subnets
                required_bits = math.ceil(math.log2(num_subnets_int))
                new_prefix = base_prefix + required_bits
                
                # Calculate how many addresses we need for all subnets
                addresses_per_subnet = 2 ** (32 - new_prefix)
                total_addresses_needed = num_subnets_int * addresses_per_subnet
                
                # Calculate the required prefix length for the expanded network (if needed)
                required_prefix = 32 - math.ceil(math.log2(total_addresses_needed))
                
                # Check if we need to expand the network
                base_addresses = 2 ** (32 - base_prefix)
                
                # Only expand if we don't have enough space
                if total_addresses_needed > base_addresses:
                    if required_prefix < base_prefix:
                        base_network_int = int(base_network.network_address)
                        block_size = 2 ** (32 - required_prefix)
                        aligned_network_int = (base_network_int // block_size) * block_size
                        if aligned_network_int > base_network_int:
                            aligned_network_int -= block_size
                        expanded_network = ipaddress.ip_network((aligned_network_int, required_prefix), strict=False)
                        base_network = expanded_network
                        base_prefix = expanded_network.prefixlen
                
                # Check if we have enough bits
                max_subnets = 2 ** (32 - base_prefix)
                if num_subnets_int > max_subnets:
                    return None, f'Cannot create {num_subnets_int} subnets from {base_subnet_str}. Maximum possible: {max_subnets}'
                
                # Generate all subnets
                all_subnets = list(base_network.subnets(new_prefix=new_prefix))
                subnets = all_subnets[:num_subnets_int]
                
                # Collect gateway IP from each subnet
                gateway_ips = []
                for subnet in subnets:
                    host_ips = list(subnet.hosts())
                    if len(host_ips) < 1:
                        continue
                    gateway_ip = host_ips[0]
                    gateway_ips.append({
                        'gateway': str(gateway_ip),
                        'subnet': subnet
                    })
                
                if len(gateway_ips) == 0:
                    return None, 'No usable host IPs in the calculated subnets'
                
                # Sort by gateway IP (numerically)
                gateway_ips.sort(key=lambda x: ipaddress.IPv4Address(x['gateway']))
                
                return gateway_ips, None
            
            # Divide service base subnet
            service_gateway_ips, error = divide_network(base_subnet, num_subnets)
            if error:
                return jsonify({'error': error}), 400
            
            # Divide OM base subnet if creating pair
            om_gateway_ips = []
            if create_pair:
                om_gateway_ips, om_error = divide_network(om_base_subnet, num_subnets)
                if om_error:
                    return jsonify({'error': f'OM subnet error: {om_error}'}), 400
                
                # Validate that both generate the same number of gateway IPs
                if len(service_gateway_ips) != len(om_gateway_ips):
                    return jsonify({
                        'error': f'Service base subnet generates {len(service_gateway_ips)} gateway IPs but OM base subnet generates {len(om_gateway_ips)} gateway IPs. Both must generate the same number.'
                    }), 400
            
            # Check for existing IPs
            existing_ips = set()
            for gw in service_gateway_ips:
                if IP.query.filter_by(gateway=gw['gateway']).first():
                    existing_ips.add(gw['gateway'])
            
            if create_pair:
                for gw in om_gateway_ips:
                    if IP.query.filter_by(gateway=gw['gateway']).first():
                        existing_ips.add(gw['gateway'])
            
            if existing_ips:
                return jsonify({
                    'error': f'Some IPs already exist: {", ".join(list(existing_ips)[:5])}{"..." if len(existing_ips) > 5 else ""}'
                }), 400
            
            # Get subnet mask from first subnet
            subnet_mask = str(service_gateway_ips[0]['subnet'].netmask)
            om_subnet_mask = str(om_gateway_ips[0]['subnet'].netmask) if create_pair and om_gateway_ips else None
            
            # Create gateway IPs with pairs
            created_ips = []
            created_om_ips = []
            for idx, service_gw in enumerate(service_gateway_ips):
                # Generate unique pair_id for each pair
                pair_id = None
                if create_pair:
                    pair_id = str(uuid.uuid4())
                
                # Create service IP
                gateway_ip_obj = IP(
                    gateway=service_gw['gateway'],
                    subnet_mask=subnet_mask,
                    type=tech_name,
                    vendor_id=vendor_obj.id if vendor_obj else None,
                    vendor=vendor_name,
                    status=StatusType.FREE,
                    pair_id=pair_id if create_pair else None,
                    pair_type='service' if create_pair else None
                )
                db.session.add(gateway_ip_obj)
                created_ips.append(gateway_ip_obj)
                
                # Create OM IP if creating pair
                if create_pair:
                    om_gw = om_gateway_ips[idx]
                    om_ip = IP(
                        gateway=om_gw['gateway'],
                        subnet_mask=om_subnet_mask,
                        type=f"{tech_name}_OM",
                        vendor_id=vendor_obj.id if vendor_obj else None,
                        vendor=vendor_name,
                        status=StatusType.FREE,
                        pair_id=pair_id,
                        pair_type='om'
                    )
                    db.session.add(om_ip)
                    created_om_ips.append(om_ip)
            
            db.session.commit()
            
            # Log activity for bulk creation
            log_activity('create_ip', 'ip', created_ips[0].id if created_ips else None, 
                        f'Created {len(created_ips)} IPs from {num_subnets} subnets (base: {base_subnet})' + (f' with {len(created_om_ips)} OM pairs' if created_om_ips else ''))
            
            result = {
                'message': f'Successfully created {len(created_ips)} IP addresses from {num_subnets} subnets' + (' (with OM pairs)' if created_om_ips else ''),
                'count': len(created_ips),
                'num_subnets': num_subnets,
                'ips': [ip.to_dict() for ip in created_ips[:10]]  # Return first 10 for preview
            }
            if created_om_ips:
                result['om_ips'] = [ip.to_dict() for ip in created_om_ips[:10]]
                result['count'] = len(created_ips) + len(created_om_ips)
            
            return jsonify(result), 201
            
        except ValueError as e:
            return jsonify({'error': f'Invalid subnet format: {str(e)}'}), 400
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': f'Error creating IPs: {str(e)}'}), 500
    
    elif add_method == 'subnet':
        # Handle subnet-based IP generation (with optional pair creation)
        subnet_cidr = data.get('subnet')
        create_pair = data.get('create_pair', False)
        om_subnet_cidr = data.get('om_subnet')  # Optional OM subnet for pair creation
        
        if not subnet_cidr:
            return jsonify({'error': 'Subnet (CIDR) is required'}), 400
        
        if create_pair and not om_subnet_cidr:
            return jsonify({'error': 'OM subnet (CIDR) is required when creating a pair'}), 400
        
        try:
            # Parse the service subnet
            network = ipaddress.ip_network(subnet_cidr, strict=False)
            subnet_mask = str(network.netmask)
            
            # Get all host IPs (exclude network and broadcast addresses)
            host_ips = list(network.hosts())
            
            if len(host_ips) == 0:
                return jsonify({'error': 'Subnet has no usable host IPs'}), 400
            
            # Parse OM subnet if creating pair
            om_host_ips = []
            om_subnet_mask = None
            if create_pair:
                om_network = ipaddress.ip_network(om_subnet_cidr, strict=False)
                om_subnet_mask = str(om_network.netmask)
                om_host_ips = list(om_network.hosts())
                
                if len(om_host_ips) == 0:
                    return jsonify({'error': 'OM subnet has no usable host IPs'}), 400
                
                # Validate that both subnets generate the same number of IPs
                if len(host_ips) != len(om_host_ips):
                    return jsonify({
                        'error': f'Service subnet generates {len(host_ips)} IPs but OM subnet generates {len(om_host_ips)} IPs. Both subnets must generate the same number of IPs.'
                    }), 400
            
            # Check for existing IPs in service subnet
            existing_ips = set()
            for ip_str in [str(ip) for ip in host_ips]:
                if IP.query.filter_by(gateway=ip_str).first():
                    existing_ips.add(ip_str)
            
            # Check for existing IPs in OM subnet if creating pair
            if create_pair:
                for ip_str in [str(ip) for ip in om_host_ips]:
                    if IP.query.filter_by(gateway=ip_str).first():
                        existing_ips.add(ip_str)
            
            if existing_ips:
                return jsonify({
                    'error': f'Some IPs already exist: {", ".join(list(existing_ips)[:5])}{"..." if len(existing_ips) > 5 else ""}'
                }), 400
            
            # Generate pair_id if creating pairs
            import uuid
            
            # Create all IPs
            created_ips = []
            created_om_ips = []
            for idx, ip_addr in enumerate(host_ips):
                # Generate unique pair_id for each pair
                pair_id = None
                if create_pair:
                    pair_id = str(uuid.uuid4())
                
                # Create service IP
                ip = IP(
                    gateway=str(ip_addr),
                    subnet_mask=subnet_mask,
                    type=tech_name,
                    vendor_id=vendor_obj.id if vendor_obj else None,
                    vendor=vendor_name,
                    status=StatusType.FREE,
                    pair_id=pair_id if create_pair else None,
                    pair_type='service' if create_pair else None
                )
                db.session.add(ip)
                created_ips.append(ip)
                
                # Create OM IP if creating pair
                if create_pair:
                    om_ip_addr = om_host_ips[idx]
                    om_ip = IP(
                        gateway=str(om_ip_addr),
                        subnet_mask=om_subnet_mask,
                        type=f"{tech_name}_OM",
                        vendor_id=vendor_obj.id if vendor_obj else None,
                        vendor=vendor_name,
                        status=StatusType.FREE,
                        pair_id=pair_id,
                        pair_type='om'
                    )
                    db.session.add(om_ip)
                    created_om_ips.append(om_ip)
            
            db.session.commit()
            
            # Log activity for bulk creation
            log_activity('create_ip', 'ip', created_ips[0].id if created_ips else None, 
                        f'Created {len(created_ips)} IPs from subnet {subnet_cidr}' + (f' with {len(created_om_ips)} OM pairs' if created_om_ips else ''))
            
            result = {
                'message': f'Successfully created {len(created_ips)} IP addresses from subnet {subnet_cidr}' + (' (with OM pairs)' if created_om_ips else ''),
                'count': len(created_ips),
                'ips': [ip.to_dict() for ip in created_ips[:10]]  # Return first 10 for preview
            }
            if created_om_ips:
                result['om_ips'] = [ip.to_dict() for ip in created_om_ips[:10]]
                result['count'] = len(created_ips) + len(created_om_ips)
            
            return jsonify(result), 201
            
        except ValueError as e:
            return jsonify({'error': f'Invalid subnet format: {str(e)}'}), 400
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': f'Error creating IPs: {str(e)}'}), 500
    
    else:
        # Handle single IP creation (with optional pair creation)
        gateway_address = data.get('gateway')
        om_gateway_address = data.get('om_gateway')  # Optional OM gateway for pair creation
        subnet_mask = data.get('subnet_mask', '')
        create_pair = data.get('create_pair', False)  # Whether to create as a pair
        
        if not gateway_address:
            return jsonify({'error': 'Gateway is required'}), 400
        
        if IP.query.filter_by(gateway=gateway_address).first():
            return jsonify({'error': 'Gateway already exists'}), 400
        
        # Generate pair_id if creating a pair
        pair_id = None
        if create_pair:
            import uuid
            pair_id = str(uuid.uuid4())
            if not om_gateway_address:
                return jsonify({'error': 'OM gateway is required when creating a pair'}), 400
            if IP.query.filter_by(gateway=om_gateway_address).first():
                return jsonify({'error': 'OM gateway already exists'}), 400
        
        # Create service IP
        service_ip = IP(
            gateway=gateway_address,
            subnet_mask=subnet_mask,
            type=tech_name,
            vendor_id=vendor_obj.id if vendor_obj else None,
            vendor=vendor_name,
            status=StatusType.FREE,
            pair_id=pair_id if create_pair else None,
            pair_type='service' if create_pair else None
        )
        db.session.add(service_ip)
        
        # Create OM IP if pair is requested
        om_ip = None
        if create_pair:
            om_ip = IP(
                gateway=om_gateway_address,
                subnet_mask=subnet_mask,
                type=f"{tech_name}_OM",
                vendor_id=vendor_obj.id if vendor_obj else None,
                vendor=vendor_name,
                status=StatusType.FREE,
                pair_id=pair_id,
                pair_type='om'
            )
            db.session.add(om_ip)
        
        db.session.commit()
        
        log_activity('create_ip', 'ip', service_ip.id, 
                    f'Created gateway {gateway_address}' + (f' with OM pair {om_gateway_address}' if create_pair else ''))
        
        result = {
            'message': 'IP created successfully' + (' (with OM pair)' if create_pair else ''),
            'ip': service_ip.to_dict()
        }
        if om_ip:
            result['om_ip'] = om_ip.to_dict()
        
        return jsonify(result), 201

@app.route('/subnet-calculator')
@login_required
def subnet_calculator():
    """Subnetting calculator page"""
    return render_template('subnet_calculator.html')

@app.route('/api/subnet-calculator', methods=['POST'])
@login_required
def api_calculate_subnets():
    """Calculate subnets from a base subnet"""
    data = request.json
    base_subnet = data.get('base_subnet')
    num_subnets = data.get('num_subnets')
    
    if not base_subnet or not num_subnets:
        return jsonify({'error': 'Base subnet and number of subnets are required'}), 400
    
    try:
        num_subnets = int(num_subnets)
        if num_subnets < 2:
            return jsonify({'error': 'Number of subnets must be at least 2'}), 400
    except (ValueError, TypeError):
        return jsonify({'error': 'Number of subnets must be a valid integer'}), 400
    
    try:
        # Parse the base network
        base_network = ipaddress.ip_network(base_subnet, strict=False)
        base_prefix = base_network.prefixlen
        
        # Calculate required prefix length for the number of subnets
        # We need 2^n >= num_subnets, so n = ceil(log2(num_subnets))
        required_bits = math.ceil(math.log2(num_subnets))
        new_prefix = base_prefix + required_bits
        
        # Check if we have enough bits
        max_subnets = 2 ** (32 - base_prefix)
        if num_subnets > max_subnets:
            return jsonify({
                'error': f'Cannot create {num_subnets} subnets from {base_subnet}. Maximum possible: {max_subnets}'
            }), 400
        
        # Generate all subnets
        subnets = list(base_network.subnets(new_prefix=new_prefix))
        
        # Limit to requested number
        subnets = subnets[:num_subnets]
        
        # Format results
        results = []
        for subnet in subnets:
            hosts = list(subnet.hosts())
            results.append({
                'network_address': str(subnet.network_address),
                'subnet_mask': str(subnet.netmask),
                'cidr_notation': f'{subnet.network_address}/{subnet.prefixlen}',
                'usable_hosts': len(hosts),
                'first_host': str(hosts[0]) if hosts else 'N/A',
                'last_host': str(hosts[-1]) if hosts else 'N/A',
                'broadcast_address': str(subnet.broadcast_address)
            })
        
        return jsonify({
            'base_subnet': base_subnet,
            'num_subnets': len(results),
            'subnet_mask': str(subnets[0].netmask) if subnets else 'N/A',
            'cidr_prefix': new_prefix,
            'hosts_per_subnet': len(list(subnets[0].hosts())) if subnets else 0,
            'subnets': results
        })
        
    except ValueError as e:
        return jsonify({'error': f'Invalid subnet format: {str(e)}'}), 400
    except Exception as e:
        return jsonify({'error': f'Error calculating subnets: {str(e)}'}), 500

@app.route('/api/ips/<int:ip_id>', methods=['DELETE'])
@login_required
@admin_required
def api_delete_ip(ip_id):
    """Delete an IP"""
    ip = IP.query.get_or_404(ip_id)
    
    # Check if IP is assigned to a site
    if ip.sites:
        return jsonify({'error': 'Cannot delete IP assigned to a site'}), 400
    
    db.session.delete(ip)
    db.session.commit()
    
    return jsonify({'message': 'IP deleted successfully'}), 200

@app.route('/api/ips/bulk-delete', methods=['POST'])
@login_required
@admin_required
def api_bulk_delete_ips():
    """Bulk delete IPs. Skips IPs assigned to sites and reports them."""
    try:
        data = request.json or {}
        ip_ids = data.get('ip_ids', [])
        if not isinstance(ip_ids, list) or len(ip_ids) == 0:
            return jsonify({'error': 'ip_ids must be a non-empty array'}), 400

        # Fetch all IPs in one query
        ips = IP.query.filter(IP.id.in_(ip_ids)).all()
        found_ids = {ip.id for ip in ips}
        missing_ids = [i for i in ip_ids if i not in found_ids]

        # Separate deletable from protected (assigned)
        deletable = [ip for ip in ips if not ip.sites]
        protected = [ip for ip in ips if ip.sites]

        # Delete in-session
        for ip in deletable:
            db.session.delete(ip)
        db.session.commit()

        return jsonify({
            'deleted_count': len(deletable),
            'skipped_assigned_count': len(protected),
            'missing_count': len(missing_ids),
            'skipped_assigned_ids': [ip.id for ip in protected],
            'missing_ids': missing_ids
        })
    except Exception as e:
        app.logger.error(f'Error in bulk delete IPs: {str(e)}', exc_info=True)
        db.session.rollback()
        return jsonify({'error': 'Failed to bulk delete IPs'}), 500
# Vendor Management Routes
@app.route('/vendors')
@login_required
def vendors():
    return render_template('vendors.html')

@app.route('/api/vendors', methods=['GET'])
@login_required
def api_get_vendors():
    """Get vendors with pagination"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    search = request.args.get('search', "", type=str)
    
    query = Vendor.query.order_by(Vendor.name)
    if search:
        query = query.filter(Vendor.name.like(f'%{search}%'))
    
    pagination = query.paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    return jsonify({
        'vendors': [v.to_dict() for v in pagination.items],
        'total': pagination.total,
        'pages': pagination.pages,
        'current_page': page,
        'search': search
    })

@app.route('/api/vendors', methods=['POST'])
@login_required
@admin_required
def api_create_vendor():
    """Create a new vendor"""
    data = request.json
    name = data.get('name')
    
    if not name:
        return jsonify({'error': 'Vendor name is required'}), 400
    
    if Vendor.query.filter_by(name=name).first():
        return jsonify({'error': 'Vendor with this name already exists'}), 400
    
    vendor = Vendor(name=name)
    db.session.add(vendor)
    db.session.commit()
    
    log_activity('create_vendor', 'vendor', vendor.id, vendor.name)
    
    return jsonify({'message': 'Vendor created successfully', 'vendor': vendor.to_dict()}), 201

@app.route('/api/vendors/<int:vendor_id>', methods=['PUT'])
@login_required
@admin_required
def api_update_vendor(vendor_id):
    """Update a vendor"""
    vendor = Vendor.query.get_or_404(vendor_id)
    data = request.json
    
    if 'name' in data:
        if Vendor.query.filter(Vendor.id != vendor_id, Vendor.name == data['name']).first():
            return jsonify({'error': 'Vendor with this name already exists'}), 400
        vendor.name = data['name']
    
    db.session.commit()
    
    return jsonify({'message': 'Vendor updated successfully', 'vendor': vendor.to_dict()}), 200

@app.route('/api/vendors/<int:vendor_id>', methods=['DELETE'])
@login_required
@admin_required
def api_delete_vendor(vendor_id):
    """Delete a vendor"""
    vendor = Vendor.query.get_or_404(vendor_id)
    
    # Check if vendor is used by any IPs, VLANs, or Sites
    if vendor.ips or vendor.vlans or vendor.sites:
        return jsonify({'error': 'Cannot delete vendor that is in use'}), 400
    
    db.session.delete(vendor)
    db.session.commit()
    
    return jsonify({'message': 'Vendor deleted successfully'}), 200

# Site configuration helpers
def generate_site_config_block(site_id: str, site_name: str, gateway: str, subnet: str, vlan: int,
                               vendor: str, technology: str, interface: str) -> str:
    """Render the site configuration block using the provided parameters."""
    relay_gateway = None
    if technology in ["OM", "4G_OM", "3G_OM", "2G_OM"]:
        vpn_instance = "LTE_OM"
        relay_gateway = "10.119.119.9" if vendor == "Nokia" else "10.119.10.4"
    elif technology in ["5G", "4G"]:
        vpn_instance = "LTE"
    elif technology in ["2G"]:
        vpn_instance = "AIBS"
    else:
        vpn_instance = technology

    cfg = []
    cfg.append(f"interface {interface}.{vlan}")
    cfg.append(f" vlan-type dot1q {vlan}")
    cfg.append(f" description {site_id}_{site_name}_{technology}_{vendor}")
    cfg.append(f" ip binding vpn-instance {vpn_instance}")
    cfg.append(f" ip address {gateway} {subnet}")
    cfg.append(f" statistic enable")
    cfg.append(f" trust upstream default")
    if relay_gateway:
        cfg.append(f" dhcp select relay")
    if relay_gateway:
        cfg.append(f" ip relay address {relay_gateway}")
    cfg.append(f" trust 8021p")
    cfg.append("#")
    return "\n".join(cfg)

    


def build_site_config_blocks(site: Site) -> list[str]:
    """Collect site details and generate configuration snippets (service and OM if available)."""
    vendor = site.vendor_obj.name if site.vendor_obj else None
    interface_name = site.interface.name if site.interface else None

    if not vendor:
        raise ValueError('Vendor is required to generate configuration')
    if not interface_name:
        raise ValueError('Interface assignment is required to generate configuration')

    configs = []

    # Service block
    if site.service_ip and site.service_vlan:
        technology = site.service_ip.type or site.technology_type
        gateway = site.service_ip.gateway
        subnet = site.service_ip.subnet_mask
        vlan = site.service_vlan.vlan_id
        if not technology or not gateway or not subnet or not vlan:
            raise ValueError(f'Missing service IP/VLAN data for site {site.site_id}')
        configs.append(generate_site_config_block(
            site_id=site.site_id,
            site_name=site.site_name,
            gateway=gateway,
            subnet=subnet,
            vlan=vlan,
            vendor=vendor,
            technology=technology,
            interface=interface_name
        ))

    # OM block (when paired)
    if site.om_ip and site.om_vlan:
        technology = site.om_ip.type or (f"{site.technology_type}_OM" if site.technology_type else None)
        gateway = site.om_ip.gateway
        subnet = site.om_ip.subnet_mask
        vlan = site.om_vlan.vlan_id
        if not technology or not gateway or not subnet or not vlan:
            raise ValueError(f'Missing OM IP/VLAN data for site {site.site_id}')
        configs.append(generate_site_config_block(
            site_id=site.site_id,
            site_name=site.site_name,
            gateway=gateway,
            subnet=subnet,
            vlan=vlan,
            vendor=vendor,
            technology=technology,
            interface=interface_name
        ))

    if not configs:
        raise ValueError(f'No IP/VLAN assignments found for site {site.site_id}')

    return configs

# Site Management Routes
@app.route('/sites')
@login_required
def sites():
    return render_template('sites.html')

@app.route('/api/sites', methods=['GET'])
@login_required
def api_get_sites():
    """Get sites with filtering and pagination"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    technology = request.args.get('technology')
    vendor = request.args.get('vendor')
    search = request.args.get('search', '')
    
    query = Site.query
    
    if technology:
        tech_name = parse_technology(technology)
        if tech_name:
            query = query.filter_by(technology_type=tech_name)

    # Vendor filter: accept vendor name (for backward compatibility) and
    # translate it to vendor_id, since Site now stores only vendor_id.
    if vendor:
        vendor_obj = Vendor.query.filter_by(name=vendor).first()
        if vendor_obj:
            query = query.filter_by(vendor_id=vendor_obj.id)
        else:
            # No sites will match an unknown vendor; force empty result.
            query = query.filter(db.text('1=0'))
    if search:
        query = query.filter(
            db.or_(
                Site.site_name.like(f'%{search}%'),
                Site.site_id.like(f'%{search}%')
            )
        )
    
    pagination = query.order_by(Site.site_id.asc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    return jsonify({
        'sites': [s.to_dict() for s in pagination.items],
        'total': pagination.total,
        'pages': pagination.pages,
        'current_page': page
    })


@app.route('/api/sites/config', methods=['POST'])
@login_required
def api_generate_site_configs():
    """Generate configuration blocks for selected sites."""
    try:
        data = request.json or {}
        site_ids = data.get('site_ids', [])

        if not site_ids or not isinstance(site_ids, list):
            return jsonify({'error': 'site_ids must be a non-empty array'}), 400

        try:
            normalized_site_ids = [int(site_id) for site_id in site_ids]
        except (TypeError, ValueError) as e:
            app.logger.warning('Invalid site_ids payload for config generation: %s', str(e))
            return jsonify({'error': 'site_ids must contain valid integers'}), 400

        sites = Site.query.filter(Site.id.in_(normalized_site_ids)).all()
        found_ids = {s.id for s in sites}
        missing_ids = [sid for sid in normalized_site_ids if sid not in found_ids]

        if missing_ids:
            app.logger.warning('Config requested for missing site IDs: %s', missing_ids)
            return jsonify({'error': f'Site IDs not found: {missing_ids}'}), 404

        configs = []
        for site in sites:
            blocks = build_site_config_blocks(site)
            for block in blocks:
                configs.append({
                    'site_id': site.site_id,
                    'site_name': site.site_name,
                    'config': block
                })

        combined_config = "\n\n".join(cfg['config'] for cfg in configs)

        return jsonify({
            'configs': configs,
            'combined_config': combined_config
        }), 200
    except ValueError as ve:
        app.logger.warning('Config generation validation error: %s', str(ve))
        return jsonify({'error': str(ve)}), 400
    except Exception as e:
        app.logger.error('Error generating site configurations: %s', str(e), exc_info=True)
        return jsonify({'error': 'Failed to generate site configurations'}), 500

@app.route('/api/sites', methods=['POST'])
@login_required
@write_access_required
def api_create_site():
    """Create a new site with automatic IP and VLAN assignment"""
    try:
        data = request.json
        site_id = data.get('site_id')
        site_name = data.get('site_name')
        technologies = data.get('technologies', [])  # Array of technologies
        vendor_id = data.get('vendor_id')
        router_id = data.get('router_id')
        interface_id = data.get('interface_id')
        
        app.logger.info(f'Creating site: site_id={site_id}, site_name={site_name}, technologies={technologies}, vendor_id={vendor_id}, router_id={router_id}, interface_id={interface_id}')
        
        # Required fields: site_id, site_name, technologies, vendor_id
        # router_id and interface_id are optional
        if not all([site_id, site_name, technologies, vendor_id]):
            app.logger.warning(f'Missing required fields for site creation: site_id={site_id}, site_name={site_name}, technologies={technologies}, vendor_id={vendor_id}')
            return jsonify({'error': 'Site ID, Site Name, Technologies, and Vendor are required'}), 400
        
        if not isinstance(technologies, list) or len(technologies) == 0:
            app.logger.warning(f'Invalid technologies list: {technologies}')
            return jsonify({'error': 'At least one technology is required'}), 400
        
        # Validate technologies
        tech_names = []
        for tech in technologies:
            tech_name = parse_technology(tech)
            if not tech_name:
                app.logger.warning(f'Invalid technology: {tech}')
                return jsonify({'error': f'Invalid technology: {tech}'}), 400
            tech_names.append(tech_name)
        
        # Note: site_id can be duplicated for different technologies, so we don't check for uniqueness here
        
        # Get vendor
        try:
            vendor = Vendor.query.get_or_404(vendor_id)
        except Exception as e:
            app.logger.error(f'Error fetching vendor {vendor_id}: {str(e)}', exc_info=True)
            return jsonify({'error': f'Vendor not found: {vendor_id}'}), 404
        
        # Get router and interface (optional)
        router = None
        interface = None
        if router_id and interface_id:
            try:
                router = Router.query.get_or_404(router_id)
                interface = Interface.query.get_or_404(interface_id)
                
                # Verify interface belongs to router
                if interface.router_id != router_id:
                    app.logger.warning(f'Interface {interface_id} does not belong to router {router_id}')
                    return jsonify({'error': 'Interface does not belong to the selected router'}), 400
                
                # Get all VLAN IDs already used on this interface (for VLAN reuse check)
                used_vlan_ids = []
                for s in Site.query.filter_by(interface_id=interface_id).all():
                    if s.service_vlan_id:
                        used_vlan_ids.append(s.service_vlan_id)
                    if s.om_vlan_id:
                        used_vlan_ids.append(s.om_vlan_id)
                app.logger.info(f'Found {len(used_vlan_ids)} VLANs already used on interface {interface_id}')
            except Exception as e:
                app.logger.error(f'Error fetching router/interface: router_id={router_id}, interface_id={interface_id}, error={str(e)}', exc_info=True)
                return jsonify({'error': f'Router or Interface not found'}), 404
        else:
            # No interface specified, so no VLAN conflict checking needed
            used_vlan_ids = []
            app.logger.info('Creating site without router/interface assignment')
        
        # Create one site record for each selected technology
        # Each technology gets its own IP pair (service + OM) and VLAN pair (service + OM)
        created_sites = []
        assigned_ips = []  # Track assigned IPs to avoid duplicates
        
        for tech_name in tech_names:
            # First try to find free service IP with pair (preferred)
            # Use case-insensitive matching for technology type to handle any case variations
            service_ip_query = IP.query.filter(
                (IP.vendor_id == vendor_id) | (IP.vendor == vendor.name),
                db.func.lower(IP.type) == tech_name.lower(),
                IP.status == StatusType.FREE,
                IP.pair_id.isnot(None),
                IP.pair_type == 'service'
            )
            # Exclude already assigned IPs
            if assigned_ips:
                service_ip_query = service_ip_query.filter(~IP.id.in_(assigned_ips))
            service_ip = service_ip_query.first()
            
            # If no paired IP found, try to find any free IP (unpaired)
            if not service_ip:
                unpaired_ip_query = IP.query.filter(
                    (IP.vendor_id == vendor_id) | (IP.vendor == vendor.name),
                    db.func.lower(IP.type) == tech_name.lower(),
                    IP.status == StatusType.FREE,
                    db.or_(
                        IP.pair_id.is_(None),
                        IP.pair_type != 'service'
                    )
                )
                # Exclude already assigned IPs
                if assigned_ips:
                    unpaired_ip_query = unpaired_ip_query.filter(~IP.id.in_(assigned_ips))
                service_ip = unpaired_ip_query.first()
            
            if not service_ip:
                # Debug: Check what IPs exist for this technology and vendor
                try:
                    all_ips_for_tech = IP.query.filter(
                        (IP.vendor_id == vendor_id) | (IP.vendor == vendor.name),
                        db.func.lower(IP.type) == tech_name.lower()
                    ).all()
                    
                    free_ips = [ip for ip in all_ips_for_tech if ip.status == StatusType.FREE]
                    
                    # Build detailed error message
                    error_msg = f'No available service IPs for technology "{tech_name}" and vendor "{vendor.name}". '
                    if len(all_ips_for_tech) == 0:
                        error_msg += f'No IPs found with type "{tech_name}" for this vendor.'
                    elif len(free_ips) == 0:
                        error_msg += f'Found {len(all_ips_for_tech)} IP(s) but all are already assigned.'
                    else:
                        error_msg += f'Found {len(free_ips)} free IP(s) but they may be excluded from assignment.'
                    
                    # Also check for similar technology names (case-insensitive)
                    similar_techs = IP.query.with_entities(IP.type).filter(
                        (IP.vendor_id == vendor_id) | (IP.vendor == vendor.name),
                        IP.type.ilike(f'%{tech_name}%')
                    ).distinct().all()
                    if similar_techs and len(similar_techs) > 0:
                        tech_list = [t[0] for t in similar_techs]
                        error_msg += f' Found similar technology names: {", ".join(tech_list)}'
                    
                    app.logger.warning(f'No available service IP for technology {tech_name}, vendor {vendor.name}: {error_msg}')
                    db.session.rollback()
                    return jsonify({'error': error_msg}), 400
                except Exception as e:
                    app.logger.error(f'Error checking for available IPs: {str(e)}', exc_info=True)
                    db.session.rollback()
                    return jsonify({'error': 'Error checking for available IPs'}), 500
            
            # Find the matching OM IP if service IP has a pair
            om_ip = None
            if service_ip.pair_id and service_ip.pair_type == 'service':
                om_ip = IP.query.filter_by(
                    pair_id=service_ip.pair_id,
                    pair_type='om',
                    status=StatusType.FREE
                ).first()
                
                if not om_ip:
                    app.logger.warning(f'No available OM IP pair for service gateway {service_ip.gateway} (technology {tech_name})')
                    db.session.rollback()
                    return jsonify({'error': f'No available OM IP pair for service gateway {service_ip.gateway} (technology {tech_name})'}), 400
            
            # Find service VLAN for this technology type and vendor
            # VLAN can be reused on different interfaces, but NOT on the same interface
            # First try to find paired VLAN (preferred)
            service_vlan_query = VLAN.query.filter(
                ((VLAN.vendor_id == vendor_id) | (VLAN.vendor == vendor.name)),
                VLAN.type == tech_name,
                VLAN.pair_id.isnot(None),
                VLAN.pair_type == 'service'
            )
            if used_vlan_ids:
                service_vlan_query = service_vlan_query.filter(~VLAN.id.in_(used_vlan_ids))
            service_vlan = service_vlan_query.first()
            
            # If no paired VLAN found, try to find any VLAN (unpaired)
            if not service_vlan:
                unpaired_vlan_query = VLAN.query.filter(
                    ((VLAN.vendor_id == vendor_id) | (VLAN.vendor == vendor.name)),
                    VLAN.type == tech_name,
                    db.or_(
                        VLAN.pair_id.is_(None),
                        VLAN.pair_type != 'service'
                    )
                )
                if used_vlan_ids:
                    unpaired_vlan_query = unpaired_vlan_query.filter(~VLAN.id.in_(used_vlan_ids))
                service_vlan = unpaired_vlan_query.first()
            
            if not service_vlan:
                error_msg = f'No available service VLANs for technology {tech_name} and vendor {vendor.name}'
                if interface_id:
                    error_msg += ' on this interface'
                app.logger.warning(f'No available service VLAN: {error_msg}')
                db.session.rollback()
                return jsonify({'error': error_msg}), 400
            
            # Find the matching OM VLAN if service VLAN has a pair
            om_vlan = None
            if service_vlan.pair_id and service_vlan.pair_type == 'service':
                om_vlan = VLAN.query.filter_by(
                    pair_id=service_vlan.pair_id,
                    pair_type='om'
                ).first()
                
                if not om_vlan:
                    app.logger.warning(f'No available OM VLAN pair for service VLAN {service_vlan.vlan_id} (technology {tech_name})')
                    db.session.rollback()
                    return jsonify({'error': f'No available OM VLAN pair for service VLAN {service_vlan.vlan_id} (technology {tech_name})'}), 400
                
                # Check if OM VLAN is already used on this interface (only if interface is specified)
                if interface_id and om_vlan.id in used_vlan_ids:
                    app.logger.warning(f'OM VLAN {om_vlan.vlan_id} is already used on interface {interface_id}')
                    db.session.rollback()
                    return jsonify({'error': f'OM VLAN {om_vlan.vlan_id} is already used on this interface'}), 400
            
            # Use the same site_id for all technologies (site_id can be duplicated for different technologies)
            # But check if this site_id + technology combination already exists on this interface (if interface is specified)
            if interface_id:
                existing_site = Site.query.filter_by(
                    site_id=site_id,
                    technology_type=tech_name,
                    interface_id=interface_id
                ).first()
                
                if existing_site:
                    app.logger.warning(f'Site ID "{site_id}" with technology "{tech_name}" already exists on interface {interface_id}')
                    db.session.rollback()
                    return jsonify({'error': f'Site ID "{site_id}" with technology "{tech_name}" already exists on this interface'}), 400
            
            # Create site for this technology with pair assignments (or unpaired if pairs not available)
            site = Site(
                site_id=site_id,  # Use the same site_id for all technologies
                site_name=site_name,
                technology_type=tech_name,
                vendor_id=vendor_id,
                interface_id=interface_id,
                # Pair fields (may be None if pairs not available)
                service_ip_id=service_ip.id,
                om_ip_id=om_ip.id if om_ip else None,
                service_vlan_id=service_vlan.id,
                om_vlan_id=om_vlan.id if om_vlan else None
            )
            db.session.add(site)
            created_sites.append(site)
            
            # Update IP status (service IP always, OM IP only if pair exists)
            service_ip.status = StatusType.ASSIGNED
            service_ip.assigned_date = datetime.utcnow()
            assigned_ips.append(service_ip.id)
            if om_ip:
                om_ip.status = StatusType.ASSIGNED
                om_ip.assigned_date = datetime.utcnow()
                assigned_ips.append(om_ip.id)
            
            # Add VLANs to used list for subsequent technologies on same interface
            used_vlan_ids.append(service_vlan.id)
            if om_vlan:
                used_vlan_ids.append(om_vlan.id)
        
        try:
            db.session.commit()
            app.logger.info(f'Successfully created {len(created_sites)} site(s) for site_id={site_id}, technologies={tech_names}')
        except Exception as e:
            app.logger.error(f'Error committing site creation to database: {str(e)}', exc_info=True)
            db.session.rollback()
            return jsonify({'error': 'Error saving site to database'}), 500
        
        # Log activity for all created sites
        for site in created_sites:
            try:
                log_activity('assign_site', 'site', site.id, site.site_name,
                            site_name=site.site_name,
                            router=router.name if router else None,
                            interface=interface.name if interface else None)
            except Exception as e:
                app.logger.error(f'Error logging activity for site {site.id}: {str(e)}', exc_info=True)
        
        return jsonify({
            'message': f'Successfully created {len(created_sites)} site(s) for {len(tech_names)} technology/technologies',
            'sites': [s.to_dict() for s in created_sites],
            'count': len(created_sites)
        }), 201
    except Exception as e:
        app.logger.error(f'Unexpected error in api_create_site: {str(e)}', exc_info=True)
        db.session.rollback()
        return jsonify({'error': 'An unexpected error occurred while creating the site'}), 500

@app.route('/api/sites/template/download', methods=['GET'])
@login_required
@admin_required
def api_download_sites_template():
    """Download Excel template for bulk site import"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sites Template"
        
        # Define headers
        headers = ['Site ID', 'Site Name', 'Technologies (comma-separated)', 'Vendor Name', 'Router Name (Optional)', 'Interface Name (Optional)']
        ws.append(headers)
        
        # Style the header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Add example row
        example_row = ['SITE001', 'Example Site', '4G,5G', 'Vendor1', 'Router1', 'GigabitEthernet0/0/1']
        ws.append(example_row)
        
        # Add instructions row
        ws.append([])
        ws.append(['Instructions:'])
        ws.append(['1. Fill in the required fields: Site ID, Site Name, Technologies, and Vendor Name (Huawei, Nokia)'])
        ws.append(['2. Technologies should be comma-separated (e.g., "4G,5G")'])
        ws.append(['3. Router Name and Interface Name are optional'])
        ws.append(['4. If Router/Interface are not provided, free IPs and VLANs will be assigned automatically'])
        ws.append(['5. Delete the example row and instructions before importing'])
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 30
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        app.logger.info('Site template downloaded by user')
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='sites_import_template.xlsx'
        )
    except Exception as e:
        app.logger.error(f'Error generating site template: {str(e)}', exc_info=True)
        return jsonify({'error': 'Error generating template'}), 500

@app.route('/api/sites/bulk-import', methods=['POST'])
@login_required
@admin_required
def api_bulk_import_sites():
    """Import multiple sites from Excel file"""
    try:
        from openpyxl import load_workbook
        
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Invalid file format. Please upload an Excel file (.xlsx or .xls)'}), 400
        
        # Load workbook
        wb = load_workbook(file, data_only=True)
        ws = wb.active
        
        # Read data (skip header row)
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return jsonify({'error': 'Template is empty. Please add site data'}), 400
        
        # Get header row
        headers = [str(cell).strip() if cell else '' for cell in rows[0]]
        
        # Find column indices
        try:
            site_id_idx = headers.index('Site ID')
            site_name_idx = headers.index('Site Name')
            technologies_idx = headers.index('Technologies (comma-separated)')
            vendor_idx = headers.index('Vendor Name')
            router_idx = headers.index('Router Name (Optional)') if 'Router Name (Optional)' in headers else -1
            interface_idx = headers.index('Interface Name (Optional)') if 'Interface Name (Optional)' in headers else -1
        except ValueError as e:
            app.logger.error(f'Missing required column in template: {str(e)}')
            return jsonify({'error': f'Invalid template format. Missing required column: {str(e)}'}), 400
        
        # VALIDATION PHASE: Validate all rows before inserting anything
        errors = []
        validated_rows = []  # Store validated row data for insertion phase
        # Track resources across all rows to prevent conflicts
        all_assigned_ips = set()  # Track IPs assigned in this import
        all_used_vlan_ids = {}  # Track VLANs by interface_id: set of VLAN IDs
        
        app.logger.info('Starting validation phase for bulk import')
        
        for row_num, row in enumerate(rows[1:], start=2):  # Start from row 2 (skip header)
            # Skip empty rows
            if not any(row):
                continue
            
            # Skip instruction rows
            if any(cell and isinstance(cell, str) and ('Instructions' in cell or 'Fill in' in cell or 'Delete' in cell) for cell in row):
                continue
            
            try:
                site_id = str(row[site_id_idx]).strip() if row[site_id_idx] else None
                site_name = str(row[site_name_idx]).strip() if row[site_name_idx] else None
                technologies_str = str(row[technologies_idx]).strip() if row[technologies_idx] else None
                vendor_name = str(row[vendor_idx]).strip() if row[vendor_idx] else None
                router_name = str(row[router_idx]).strip() if row[router_idx] and router_idx >= 0 else None
                interface_name = str(row[interface_idx]).strip() if row[interface_idx] and interface_idx >= 0 else None
                
                # Validate required fields
                if not site_id or not site_name or not technologies_str or not vendor_name:
                    errors.append(f'Row {row_num}: Missing required fields')
                    continue
                
                # Parse technologies
                technologies = [t.strip() for t in technologies_str.split(',') if t.strip()]
                if not technologies:
                    errors.append(f'Row {row_num}: No valid technologies provided')
                    continue
                
                # Get vendor
                vendor = Vendor.query.filter_by(name=vendor_name).first()
                if not vendor:
                    errors.append(f'Row {row_num}: Vendor "{vendor_name}" not found')
                    continue
                
                # Get router and interface if provided
                router_id = None
                interface_id = None
                router = None
                interface = None
                if router_name and interface_name:
                    router = Router.query.filter_by(name=router_name).first()
                    if not router:
                        errors.append(f'Row {row_num}: Router "{router_name}" not found')
                        continue
                    
                    interface = Interface.query.filter_by(name=interface_name, router_id=router.id).first()
                    if not interface:
                        errors.append(f'Row {row_num}: Interface "{interface_name}" not found on router "{router_name}"')
                        continue
                    
                    router_id = router.id
                    interface_id = interface.id
                elif router_name or interface_name:
                    errors.append(f'Row {row_num}: Both Router and Interface must be provided together, or leave both empty')
                    continue
                
                # Validate technologies
                tech_names = []
                for tech in technologies:
                    tech_name = parse_technology(tech)
                    if not tech_name:
                        errors.append(f'Row {row_num}: Invalid technology "{tech}"')
                        break
                    tech_names.append(tech_name)
                
                if len(tech_names) != len(technologies):
                    continue
                
                # Get used VLAN IDs if interface is specified (from existing sites)
                existing_used_vlan_ids = []
                if interface_id:
                    for s in Site.query.filter_by(interface_id=interface_id).all():
                        if s.service_vlan_id:
                            existing_used_vlan_ids.append(s.service_vlan_id)
                        if s.om_vlan_id:
                            existing_used_vlan_ids.append(s.om_vlan_id)
                
                # Get VLANs already used in this import for this interface
                import_used_vlan_ids = all_used_vlan_ids.get(interface_id, set()) if interface_id else set()
                # Combine existing and import-used VLANs
                row_used_vlan_ids = set(existing_used_vlan_ids) | import_used_vlan_ids
                
                # Track IPs used in this row to avoid duplicates within the row
                row_assigned_ips = []
                
                # Validate each technology for this row
                row_sites_data = []
                for tech_name in tech_names:
                    # Find free service IP (not used in this import or this row)
                    service_ip_query = IP.query.filter(
                        (IP.vendor_id == vendor.id) | (IP.vendor == vendor.name),
                        db.func.lower(IP.type) == tech_name.lower(),
                        IP.status == StatusType.FREE,
                        IP.pair_id.isnot(None),
                        IP.pair_type == 'service'
                    )
                    # Exclude IPs already assigned in this import or this row
                    excluded_ips = list(all_assigned_ips | set(row_assigned_ips))
                    if excluded_ips:
                        service_ip_query = service_ip_query.filter(~IP.id.in_(excluded_ips))
                    service_ip = service_ip_query.first()
                    
                    if not service_ip:
                        unpaired_ip_query = IP.query.filter(
                            (IP.vendor_id == vendor.id) | (IP.vendor == vendor.name),
                            db.func.lower(IP.type) == tech_name.lower(),
                            IP.status == StatusType.FREE,
                            db.or_(
                                IP.pair_id.is_(None),
                                IP.pair_type != 'service'
                            )
                        )
                        excluded_ips = list(all_assigned_ips | set(row_assigned_ips))
                        if excluded_ips:
                            unpaired_ip_query = unpaired_ip_query.filter(~IP.id.in_(excluded_ips))
                        service_ip = unpaired_ip_query.first()
                    
                    if not service_ip:
                        errors.append(f'Row {row_num}: No available service IPs for technology "{tech_name}" and vendor "{vendor.name}"')
                        break
                    
                    # Find OM IP (check if not already assigned in this import)
                    om_ip = None
                    if service_ip.pair_id and service_ip.pair_type == 'service':
                        om_ip_query = IP.query.filter_by(
                            pair_id=service_ip.pair_id,
                            pair_type='om',
                            status=StatusType.FREE
                        )
                        # Exclude if already assigned in this import
                        if all_assigned_ips:
                            om_ip_query = om_ip_query.filter(~IP.id.in_(list(all_assigned_ips)))
                        om_ip = om_ip_query.first()
                        
                        if not om_ip:
                            errors.append(f'Row {row_num}: No available OM IP pair for technology "{tech_name}"')
                            break
                    
                    # Find service VLAN (not used on this interface in existing sites or this import)
                    service_vlan_query = VLAN.query.filter(
                        ((VLAN.vendor_id == vendor.id) | (VLAN.vendor == vendor.name)),
                        VLAN.type == tech_name,
                        VLAN.pair_id.isnot(None),
                        VLAN.pair_type == 'service'
                    )
                    if row_used_vlan_ids:
                        service_vlan_query = service_vlan_query.filter(~VLAN.id.in_(list(row_used_vlan_ids)))
                    service_vlan = service_vlan_query.first()
                    
                    if not service_vlan:
                        unpaired_vlan_query = VLAN.query.filter(
                            ((VLAN.vendor_id == vendor.id) | (VLAN.vendor == vendor.name)),
                            VLAN.type == tech_name,
                            db.or_(
                                VLAN.pair_id.is_(None),
                                VLAN.pair_type != 'service'
                            )
                        )
                        if row_used_vlan_ids:
                            unpaired_vlan_query = unpaired_vlan_query.filter(~VLAN.id.in_(list(row_used_vlan_ids)))
                        service_vlan = unpaired_vlan_query.first()
                    
                    if not service_vlan:
                        errors.append(f'Row {row_num}: No available service VLANs for technology "{tech_name}" and vendor "{vendor.name}"')
                        break
                    
                    # Find OM VLAN
                    om_vlan = None
                    if service_vlan.pair_id and service_vlan.pair_type == 'service':
                        om_vlan = VLAN.query.filter_by(
                            pair_id=service_vlan.pair_id,
                            pair_type='om'
                        ).first()
                        
                        if not om_vlan:
                            errors.append(f'Row {row_num}: No available OM VLAN pair for technology "{tech_name}"')
                            break
                        
                        if interface_id and om_vlan.id in row_used_vlan_ids:
                            errors.append(f'Row {row_num}: OM VLAN {om_vlan.vlan_id} is already used on this interface')
                            break
                    
                    # Check for existing site
                    if interface_id:
                        existing_site = Site.query.filter_by(
                            site_id=site_id,
                            technology_type=tech_name,
                            interface_id=interface_id
                        ).first()
                        
                        if existing_site:
                            errors.append(f'Row {row_num}: Site ID "{site_id}" with technology "{tech_name}" already exists on this interface')
                            break
                    
                    # Store validated data for insertion phase
                    row_sites_data.append({
                        'site_id': site_id,
                        'site_name': site_name,
                        'tech_name': tech_name,
                        'vendor': vendor,
                        'router': router,
                        'interface': interface,
                        'interface_id': interface_id,
                        'service_ip': service_ip,
                        'om_ip': om_ip,
                        'service_vlan': service_vlan,
                        'om_vlan': om_vlan
                    })
                    
                    # Track resources for this row to avoid duplicates
                    row_assigned_ips.append(service_ip.id)
                    if om_ip:
                        row_assigned_ips.append(om_ip.id)
                    row_used_vlan_ids.add(service_vlan.id)
                    if om_vlan:
                        row_used_vlan_ids.add(om_vlan.id)
                
                # Only add to validated_rows if all technologies passed validation
                if len(row_sites_data) == len(tech_names):
                    validated_rows.extend(row_sites_data)
                    # Update global tracking sets for cross-row conflict prevention
                    all_assigned_ips.update(row_assigned_ips)
                    if interface_id:
                        if interface_id not in all_used_vlan_ids:
                            all_used_vlan_ids[interface_id] = set()
                        all_used_vlan_ids[interface_id].update(row_used_vlan_ids)
            
            except Exception as e:
                app.logger.error(f'Error validating row {row_num}: {str(e)}', exc_info=True)
                errors.append(f'Row {row_num}: {str(e)}')
                continue
        
        # If there are any validation errors, return early without inserting anything
        if errors:
            app.logger.warning(f'Bulk import validation failed with {len(errors)} error(s). No sites will be inserted.')
            return jsonify({
                'error': 'Validation failed. Please fix the errors and try again.',
                'errors': errors[:100],  # Limit to first 100 errors
                'error_count': len(errors)
            }), 400
        
        # INSERTION PHASE: Only proceed if validation passed
        if not validated_rows:
            return jsonify({'error': 'No valid sites to import'}), 400
        
        app.logger.info(f'Validation passed. Proceeding to insert {len(validated_rows)} site(s)')
        
        created_count = 0
        created_sites = []  # Collect sites for activity logging after commit
        
        try:
            for site_data in validated_rows:
                # Create site
                site = Site(
                    site_id=site_data['site_id'],
                    site_name=site_data['site_name'],
                    technology_type=site_data['tech_name'],
                    vendor_id=site_data['vendor'].id,
                    interface_id=site_data['interface_id'],
                    service_ip_id=site_data['service_ip'].id,
                    om_ip_id=site_data['om_ip'].id if site_data['om_ip'] else None,
                    service_vlan_id=site_data['service_vlan'].id,
                    om_vlan_id=site_data['om_vlan'].id if site_data['om_vlan'] else None
                )
                db.session.add(site)
                created_count += 1
                created_sites.append((site, site_data['router'], site_data['interface']))
                
                # Update IP status
                service_ip = site_data['service_ip']
                service_ip.status = StatusType.ASSIGNED
                service_ip.assigned_date = datetime.utcnow()
                
                if site_data['om_ip']:
                    om_ip = site_data['om_ip']
                    om_ip.status = StatusType.ASSIGNED
                    om_ip.assigned_date = datetime.utcnow()
            
            # Commit all changes
            db.session.commit()
            app.logger.info(f'Bulk import completed successfully: {created_count} site(s) created')
            
        except Exception as e:
            app.logger.error(f'Error inserting sites during bulk import: {str(e)}', exc_info=True)
            db.session.rollback()
            return jsonify({'error': f'Error saving sites to database: {str(e)}'}), 500
        
        # Log activity for all created sites (after commit so they have IDs)
        for site, router, interface in created_sites:
            try:
                log_activity('assign_site', 'site', site.id, site.site_name,
                            site_name=site.site_name,
                            router=router.name if router else None,
                            interface=interface.name if interface else None)
            except Exception as e:
                app.logger.error(f'Error logging activity for site {site.id}: {str(e)}', exc_info=True)
        
        result = {
            'message': f'Import completed successfully: {created_count} site(s) created',
            'created_count': created_count,
            'error_count': 0
        }
        
        return jsonify(result), 200
        
    except Exception as e:
        app.logger.error(f'Error in bulk import: {str(e)}', exc_info=True)
        db.session.rollback()
        return jsonify({'error': f'Error processing file: {str(e)}'}), 500

@app.route('/api/sites/<int:site_id>/release', methods=['POST'])
@login_required
@admin_required
def api_release_site(site_id):
    """Release a single site (free up IPs, but VLANs can be reused)"""
    site = Site.query.get_or_404(site_id)
    
    # Free up IPs (both service and OM if using pairs)
    if site.service_ip_id:
        service_ip = IP.query.get(site.service_ip_id)
        if service_ip:
            service_ip.status = StatusType.FREE
            service_ip.assigned_date = None
    if site.om_ip_id:
        om_ip = IP.query.get(site.om_ip_id)
        if om_ip:
            om_ip.status = StatusType.FREE
            om_ip.assigned_date = None
    # VLANs can be reused, so don't update their status
    
    # Log before deleting
    log_activity('release_site', 'site', site.id, site.site_id,
                site_name=site.site_name,
                router=site.interface.router.name if site.interface else None,
                interface=site.interface.name if site.interface else None)
    
    # Delete site
    db.session.delete(site)
    db.session.commit()
    
    return jsonify({'message': 'Site released successfully'}), 200

@app.route('/api/sites/release', methods=['POST'])
@login_required
@admin_required
def api_release_sites():
    """Release one or multiple sites (free up IPs, but VLANs can be reused)"""
    data = request.json
    site_ids = data.get('site_ids', [])
    
    if not site_ids or not isinstance(site_ids, list) or len(site_ids) == 0:
        return jsonify({'error': 'At least one site ID is required'}), 400
    
    # Get all sites to release
    sites = Site.query.filter(Site.id.in_(site_ids)).all()
    
    if len(sites) != len(site_ids):
        return jsonify({'error': 'One or more site IDs are invalid'}), 400
    
    released_sites = []
    
    for site in sites:
        # Free up IPs (both service and OM if using pairs)
        if site.service_ip_id:
            service_ip = IP.query.get(site.service_ip_id)
            if service_ip:
                service_ip.status = StatusType.FREE
                service_ip.assigned_date = None
        if site.om_ip_id:
            om_ip = IP.query.get(site.om_ip_id)
            if om_ip:
                om_ip.status = StatusType.FREE
                om_ip.assigned_date = None
        # VLANs can be reused, so don't update their status
        
        # Log before deleting
        log_activity('release_site', 'site', site.id, site.site_id,
                    site_name=site.site_name,
                    router=site.interface.router.name if site.interface else None,
                    interface=site.interface.name if site.interface else None)
        
        released_sites.append(site)
        db.session.delete(site)
    
    db.session.commit()
    
    return jsonify({
        'message': f'Successfully released {len(released_sites)} site(s)',
        'count': len(released_sites)
    }), 200

@app.route('/api/sites/transfer/check', methods=['POST'])
@login_required
@write_access_required
def api_transfer_sites_check():
    """Check for VLAN conflicts before transferring sites"""
    data = request.json
    site_ids = data.get('site_ids', [])
    interface_id = data.get('interface_id')
    
    if not site_ids or not isinstance(site_ids, list) or len(site_ids) == 0:
        return jsonify({'error': 'At least one site ID is required'}), 400
    
    if not interface_id:
        return jsonify({'error': 'Interface ID is required'}), 400
    
    # Get all sites to transfer
    sites = Site.query.filter(Site.id.in_(site_ids)).all()
    
    if len(sites) != len(site_ids):
        return jsonify({'error': 'One or more site IDs are invalid'}), 400
    
    # Get all VLAN IDs already used on the new interface (both service and OM)
    used_vlan_ids = []
    for s in Site.query.filter_by(interface_id=interface_id).all():
        if s.service_vlan_id:
            used_vlan_ids.append(s.service_vlan_id)
        if s.om_vlan_id:
            used_vlan_ids.append(s.om_vlan_id)
    
    conflicts = []
    for site in sites:
        # Check for conflicts with service VLAN or OM VLAN
        has_conflict = False
        conflict_vlans = []
        
        if site.service_vlan_id and site.service_vlan_id in used_vlan_ids:
            has_conflict = True
            if site.service_vlan:
                conflict_vlans.append(f"Service VLAN {site.service_vlan.vlan_id}")
        if site.om_vlan_id and site.om_vlan_id in used_vlan_ids:
            has_conflict = True
            if site.om_vlan:
                conflict_vlans.append(f"OM VLAN {site.om_vlan.vlan_id}")
        
        if has_conflict:
            conflicts.append({
                'site_id': site.site_id,
                'site_name': site.site_name,
                'current_vlan': ', '.join(conflict_vlans) if conflict_vlans else 'N/A',
                'technology': site.technology_type if site.technology_type else None,
                'vendor': site.vendor_obj.name if site.vendor_obj else None
            })
    
    return jsonify({
        'has_conflicts': len(conflicts) > 0,
        'conflicts': conflicts,
        'count': len(conflicts)
    }), 200

@app.route('/api/sites/transfer', methods=['POST'])
@login_required
@write_access_required
def api_transfer_sites():
    """Transfer one or multiple sites to a new router and interface"""
    data = request.json
    site_ids = data.get('site_ids', [])
    router_id = data.get('router_id')
    interface_id = data.get('interface_id')
    reassign_vlans = data.get('reassign_vlans', False)  # Flag to reassign VLANs if conflicts exist
    
    if not site_ids or not isinstance(site_ids, list) or len(site_ids) == 0:
        return jsonify({'error': 'At least one site ID is required'}), 400
    
    if not router_id or not interface_id:
        return jsonify({'error': 'Router ID and Interface ID are required'}), 400
    
    # Get router and interface
    router = Router.query.get_or_404(router_id)
    interface = Interface.query.get_or_404(interface_id)
    
    # Verify interface belongs to router
    if interface.router_id != router_id:
        return jsonify({'error': 'Interface does not belong to the selected router'}), 400
    
    # Get all sites to transfer
    sites = Site.query.filter(Site.id.in_(site_ids)).all()
    
    if len(sites) != len(site_ids):
        return jsonify({'error': 'One or more site IDs are invalid'}), 400
    
    # Get all VLAN IDs already used on the new interface
    used_vlan_ids = []
    for s in Site.query.filter_by(interface_id=interface_id).all():
        if s.service_vlan_id:
            used_vlan_ids.append(s.service_vlan_id)
        if s.om_vlan_id:
            used_vlan_ids.append(s.om_vlan_id)
    
    transferred_sites = []
    
    for site in sites:
        # Store old router and interface info for logging
        old_router = site.interface.router.name if site.interface and site.interface.router else None
        old_interface = site.interface.name if site.interface else None
        
        # Check if VLANs are already used on the new interface (check both service and OM VLANs)
        vlan_conflict = False
        if site.service_vlan_id and site.service_vlan_id in used_vlan_ids:
            vlan_conflict = True
        if site.om_vlan_id and site.om_vlan_id in used_vlan_ids:
            vlan_conflict = True
        
        if vlan_conflict:
            if not reassign_vlans:
                db.session.rollback()
                return jsonify({
                    'error': f'VLAN conflict detected for site {site.site_id}. Please confirm to reassign VLANs.'
                }), 400
            
            # Need to find new VLAN pairs for this site
            # Get vendor and technology from site
            vendor_id = site.vendor_id
            vendor_name = site.vendor_obj.name if site.vendor_obj else None
            tech_name = site.technology_type
            
            if not tech_name:
                db.session.rollback()
                return jsonify({'error': f'Site {site.site_id} does not have a technology type'}), 400
            
            # Find a new service VLAN pair that matches the criteria and is NOT already used on this interface
            if used_vlan_ids:
                new_service_vlan = VLAN.query.filter(
                    ((VLAN.vendor_id == vendor_id) | (VLAN.vendor == vendor_name)),
                    VLAN.type == tech_name,
                    VLAN.pair_id.isnot(None),
                    VLAN.pair_type == 'service',
                    ~VLAN.id.in_(used_vlan_ids)
                ).first()
            else:
                # No VLANs used on this interface yet, find any service VLAN matching the criteria
                new_service_vlan = VLAN.query.filter(
                    ((VLAN.vendor_id == vendor_id) | (VLAN.vendor == vendor_name)),
                    VLAN.type == tech_name,
                    VLAN.pair_id.isnot(None),
                    VLAN.pair_type == 'service'
                ).first()
            
            if not new_service_vlan:
                db.session.rollback()
                return jsonify({
                    'error': f'No available service VLAN pairs for technology {tech_name} and vendor {vendor_name or "N/A"} on the new interface'
                }), 400
            
            # Find the matching OM VLAN
            new_om_vlan = VLAN.query.filter_by(
                pair_id=new_service_vlan.pair_id,
                pair_type='om'
            ).first()
            
            if not new_om_vlan:
                db.session.rollback()
                return jsonify({
                    'error': f'No available OM VLAN pair for service VLAN {new_service_vlan.vlan_id} (technology {tech_name})'
                }), 400
            
            # Check if OM VLAN is already used
            if new_om_vlan.id in used_vlan_ids:
                db.session.rollback()
                return jsonify({
                    'error': f'OM VLAN {new_om_vlan.vlan_id} is already used on the new interface'
                }), 400
            
            # Update site with new VLAN pairs
            site.service_vlan_id = new_service_vlan.id
            site.om_vlan_id = new_om_vlan.id
            used_vlan_ids.extend([new_service_vlan.id, new_om_vlan.id])
        else:
            # VLANs are not used on the new interface, can reuse them
            if site.service_vlan_id:
                used_vlan_ids.append(site.service_vlan_id)
            if site.om_vlan_id:
                used_vlan_ids.append(site.om_vlan_id)
        
        # Update router and interface
        site.interface_id = interface_id
        
        # Log activity (include old router/interface in resource_value)
        transfer_info = f"{site.site_name} (from {old_router}/{old_interface} to {router.name}/{interface.name})"
        log_activity('transfer_site', 'site', site.id, transfer_info,
                    site_name=site.site_name,
                    router=router.name,
                    interface=interface.name)
        
        transferred_sites.append(site)
    
    db.session.commit()
    
    return jsonify({
        'message': f'Successfully transferred {len(transferred_sites)} site(s)',
        'count': len(transferred_sites),
        'sites': [s.to_dict() for s in transferred_sites]
    }), 200

# Removed - VLANs are no longer tied to interfaces

@app.route('/api/ips/available', methods=['GET'])
@login_required
def api_get_available_ips():
    """Get available IPs for technology and vendor"""
    technology = request.args.get('technology')
    vendor = request.args.get('vendor')
    vendor_id = request.args.get('vendor_id')
    
    if not technology:
        return jsonify({'error': 'Technology is required'}), 400
    
    if not vendor_id and not vendor:
        return jsonify({'error': 'Vendor is required'}), 400
    
    tech_enum = parse_technology(technology)
    if not tech_enum:
        return jsonify({'error': 'Invalid technology'}), 400
    
    # Support both vendor_id and vendor string for backward compatibility
    if vendor_id:
        ips = IP.query.filter_by(
            type=tech_enum,
            vendor_id=int(vendor_id),
            status=StatusType.FREE
        ).all()
    else:
        ips = IP.query.filter_by(
            type=tech_enum,
            vendor=vendor,
            status=StatusType.FREE
        ).all()
    
    # Sort IPs numerically by gateway
    ips.sort(key=lambda ip: ipaddress.IPv4Address(ip.gateway))
    
    return jsonify({
        'ips': [ip.to_dict() for ip in ips],
        'count': len(ips)
    })

# User Management Routes
@app.route('/users')
@login_required
@admin_required
def users():
    return render_template('users.html')

@app.route('/api/users', methods=['GET'])
@login_required
@admin_required
def api_get_users():
    """Get users with pagination"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    
    query = User.query.order_by(User.username)
    
    pagination = query.paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    return jsonify({
        'users': [{
            'id': u.id,
            'username': u.username,
            'role': u.role.value if u.role else None,
            'created_at': u.created_at.isoformat() if u.created_at else None
        } for u in pagination.items],
        'total': pagination.total,
        'pages': pagination.pages,
        'current_page': page
    })

@app.route('/api/users', methods=['POST'])
@login_required
@admin_required
def api_create_user():
    """Create a new user"""
    data = request.json
    username = data.get('username')
    password = data.get('password')
    role = data.get('role', 'engineer')
    
    if not username or not password:
        return jsonify({'error': 'Username and password are required'}), 400
    
    if User.query.filter_by(username=username).first():
        return jsonify({'error': 'Username already exists'}), 400
    
    try:
        user_role = UserRole[role.upper()]
    except KeyError:
        return jsonify({'error': 'Invalid role'}), 400
    
    user = User(username=username, role=user_role)
    user.set_password(password)
    db.session.add(user)
    db.session.commit()
    # Create password state forcing change on first login
    try:
        state = PasswordState(user_id=user.id, must_change=True)
        db.session.add(state)
        db.session.commit()
    except Exception as e:
        app.logger.error(f'Error creating PasswordState for user {username}: {str(e)}', exc_info=True)
    
    return jsonify({'message': 'User created successfully'}), 201

@app.route('/api/users/<int:user_id>', methods=['PUT'])
@login_required
@admin_required
def api_update_user(user_id):
    """Update a user's role (and optionally password)"""
    user = User.query.get_or_404(user_id)
    data = request.json or {}
    try:
        # Update role
        if 'role' in data and data['role']:
            try:
                new_role = UserRole[data['role'].upper()]
            except KeyError:
                return jsonify({'error': 'Invalid role'}), 400
            # Prevent demoting your own admin role
            if user.id == current_user.id and new_role != UserRole.ADMIN:
                return jsonify({'error': 'You cannot change your own role from admin'}), 400
            user.role = new_role
        # Update password (optional)
        if 'password' in data and data['password']:
            if len(data['password']) < 6:
                return jsonify({'error': 'Password must be at least 6 characters long'}), 400
            user.set_password(data['password'])
        db.session.commit()
        return jsonify({'message': 'User updated successfully'}), 200
    except Exception as e:
        app.logger.error(f'Error updating user {user_id}: {str(e)}', exc_info=True)
        db.session.rollback()
        return jsonify({'error': 'Failed to update user'}), 500

@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@login_required
@admin_required
def api_delete_user(user_id):
    """Delete a user"""
    if user_id == current_user.id:
        return jsonify({'error': 'Cannot delete your own account'}), 400
    
    try:
        user = User.query.get_or_404(user_id)
        # Remove dependent rows that enforce FK constraints
        try:
            # Delete password state first (FK is NOT NULL and unique)
            state = PasswordState.query.filter_by(user_id=user_id).first()
            if state:
                db.session.delete(state)
        except Exception as e:
            app.logger.error(f'Error deleting password state for user {user.username}: {str(e)}', exc_info=True)
            db.session.rollback()
            return jsonify({'error': 'Failed to delete user password state'}), 500
        
        try:
            # Delete activity logs for this user to avoid FK violations
            ActivityLog.query.filter_by(user_id=user_id).delete(synchronize_session=False)
        except Exception as e:
            app.logger.error(f'Error deleting activity logs for user {user.username}: {str(e)}', exc_info=True)
            db.session.rollback()
            return jsonify({'error': 'Failed to delete user activity logs'}), 500
        
        db.session.delete(user)
        db.session.commit()
    except Exception as e:
        app.logger.error(f'Error deleting user {user_id}: {str(e)}', exc_info=True)
        db.session.rollback()
        return jsonify({'error': 'Failed to delete user'}), 500
    
    return jsonify({'message': 'User deleted successfully'}), 200

# Change own password (forced on first login or optional later)
@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        try:
            current_pwd = request.form.get('current_password')
            new_pwd = request.form.get('new_password')
            confirm_pwd = request.form.get('confirm_password')

            if not new_pwd or not confirm_pwd:
                flash('New password and confirmation are required', 'error')
                return redirect(url_for('change_password'))
            if new_pwd != confirm_pwd:
                flash('New passwords do not match', 'error')
                return redirect(url_for('change_password'))
            if len(new_pwd) < 6:
                flash('New password must be at least 6 characters long', 'error')
                return redirect(url_for('change_password'))

            # If state requires change, allow skipping current check for admin reset cases
            state = PasswordState.query.filter_by(user_id=current_user.id).first()
            require_current = not (state and state.must_change)

            if require_current:
                if not current_pwd or not current_user.check_password(current_pwd):
                    flash('Current password is incorrect', 'error')
                    return redirect(url_for('change_password'))

            current_user.set_password(new_pwd)
            # update state
            if not state:
                state = PasswordState(user_id=current_user.id, must_change=False, last_changed=datetime.utcnow())
                db.session.add(state)
            else:
                state.must_change = False
                state.last_changed = datetime.utcnow()
            db.session.commit()
            flash('Password changed successfully', 'success')
            return redirect(url_for('dashboard'))
        except Exception as e:
            app.logger.error(f'Error changing password: {str(e)}', exc_info=True)
            db.session.rollback()
            flash('Failed to change password', 'error')
            return redirect(url_for('change_password'))

    # GET
    try:
        state = PasswordState.query.filter_by(user_id=current_user.id).first()
    except Exception:
        state = None
    return render_template('change_password.html', must_change=bool(state and state.must_change))

# Admin: reset another user's password and require change on next login
@app.route('/api/users/<int:user_id>/reset-password', methods=['PUT'])
@login_required
@admin_required
def api_admin_reset_password(user_id):
    if user_id == current_user.id:
        return jsonify({'error': 'Use the change password page to change your own password'}), 400
    user = User.query.get_or_404(user_id)
    data = request.json or {}
    new_password = data.get('new_password')
    if not new_password or len(new_password) < 6:
        return jsonify({'error': 'New password must be at least 6 characters long'}), 400
    try:
        user.set_password(new_password)
        state = PasswordState.query.filter_by(user_id=user.id).first()
        if not state:
            state = PasswordState(user_id=user.id, must_change=True, last_changed=None)
            db.session.add(state)
        else:
            state.must_change = True
            state.last_changed = None
        db.session.commit()
        return jsonify({'message': 'Password reset successfully; user must change it at next login'}), 200
    except Exception as e:
        app.logger.error(f'Error resetting password for user {user.username}: {str(e)}', exc_info=True)
        db.session.rollback()
        return jsonify({'error': 'Failed to reset password'}), 500

@app.route('/api/activity-logs', methods=['GET'])
@login_required
def api_get_activity_logs():
    """Get activity logs"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    
    query = ActivityLog.query
    
    if not current_user.is_admin():
        query = query.filter_by(user_id=current_user.id)
    
    pagination = query.order_by(ActivityLog.timestamp.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    return jsonify({
        'logs': [log.to_dict() for log in pagination.items],
        'total': pagination.total,
        'pages': pagination.pages,
        'current_page': page
    })

# CSV Export/Import Routes
@app.route('/api/export/sites', methods=['GET'])
@login_required
@admin_required
def api_export_sites():
    """Export sites to CSV"""
    sites = Site.query.all()
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'Site ID', 'Site Name', 'Technologies', 'Vendor',
        'Service GW IP', 'Service Subnet Mask', 'Service VLAN',
        'OM GW IP', 'OM Subnet Mask', 'OM VLAN',
        'Router', 'Router IP', 'Interface', 'Assigned Date'
    ])
    
    for site in sites:
        # Each site now has a single technology_type; expose it as a one-element list
        tech_list = [site.technology_type] if site.technology_type else []

        service_ip_obj = getattr(site, 'service_ip', None)
        om_ip_obj = getattr(site, 'om_ip', None)

        service_ip = service_ip_obj.gateway if service_ip_obj else ''
        service_mask = service_ip_obj.subnet_mask if service_ip_obj and service_ip_obj.subnet_mask else ''

        om_ip = om_ip_obj.gateway if om_ip_obj else ''
        om_mask = om_ip_obj.subnet_mask if om_ip_obj and om_ip_obj.subnet_mask else ''

        service_vlan = site.service_vlan.vlan_id if getattr(site, 'service_vlan', None) else ''
        om_vlan = site.om_vlan.vlan_id if getattr(site, 'om_vlan', None) else ''

        writer.writerow([
            site.site_id,
            site.site_name,
            ', '.join(tech_list),
            site.vendor_obj.name if site.vendor_obj else '',
            service_ip,
            service_mask,
            service_vlan,
            om_ip,
            om_mask,
            om_vlan,
            site.interface.router.name if site.interface and site.interface.router else '',
            site.interface.router.router_ip if site.interface and site.interface.router else '',
            site.interface.name if site.interface else '',
            site.assigned_date.isoformat() if site.assigned_date else ''
        ])
    
    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8')),
        mimetype='text/csv',
        as_attachment=True,
        download_name='sites_export.csv'
    )

@app.route('/api/stats', methods=['GET'])
@login_required
def api_get_stats():
    """Get dashboard statistics"""
    total_routers = Router.query.count()
    total_interfaces = Interface.query.count()
    total_sites = Site.query.count()
    total_ips = IP.query.count()
    assigned_ips = IP.query.filter_by(status=StatusType.ASSIGNED).count()
    free_ips = total_ips - assigned_ips
    total_vlans = VLAN.query.count()
    assigned_vlans = VLAN.query.filter_by(status=StatusType.ASSIGNED).count()
    free_vlans = total_vlans - assigned_vlans
    
    return jsonify({
        'routers': total_routers,
        'interfaces': total_interfaces,
        'sites': total_sites,
        'ips': {
            'total': total_ips,
            'assigned': assigned_ips,
            'free': free_ips
        },
        'vlans': {
            'total': total_vlans,
            'assigned': assigned_vlans,
            'free': free_vlans
        }
    })

# Initialize database
def init_db():
    """Initialize database and create default admin user"""
    with app.app_context():
        db.create_all()
        # Create default admin user if it doesn't exist
        if not User.query.filter_by(role=UserRole.ADMIN).first():
            admin = User(username='admin', role=UserRole.ADMIN)
            admin.set_password('admin')
            db.session.add(admin)
            db.session.commit()
            try:
                state = PasswordState(user_id=admin.id, must_change=True)
                db.session.add(state)
                db.session.commit()
            except Exception as e:
                app.logger.error(f'Error creating password state for default admin: {str(e)}', exc_info=True)




if __name__ == '__main__':
    init_db()

    if Config.DEBUG:
        app.run(host=Config.HOST, port=Config.PORT, debug=Config.DEBUG)
    else:
        serve(app, host=Config.HOST, port=Config.PORT)
